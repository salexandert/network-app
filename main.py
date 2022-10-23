"""
    .Synopsis
    Builds PepsiCo IP Address Spreadsheet based on input parameters on summary sheet. 
    Ask Stephen Twait for additional information.  
    .Description
    See Synopsis

"""
import ipaddress
import os
import re
from colorama import Fore, Style, init
from nornir import InitNornir
from nornir.core.filter import F
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import column_index_from_string
import inspect
import collections
from ruamel.yaml import YAML
from constants import NORMALIZED_INTERFACES, INTERFACE_NAME_RE
from utils import write_hosts_data_to_inventory_file, normalize_interface_name, file_check, update_interfaces_from_config_files, check_hosts_in_inventory, directory_check
import pandas as pd
from time import strftime

# init() Used to fix colorama issues
init()

def generate_networks_from_ffa(summary_data):
    """This will use the FFA Block found on the Summary sheet to Populate Reccomended IP Networks. 

    Args:
        filename ([str]): [Pepsi_Paramaters.xlsx]
        checked (bool, optional): [Don't check filename again]. Defaults to False.
    """   
    
    # load the excel document
    ffa_supernet = ipaddress.ip_network(summary_data['ffa_block_xls'])
    
    # # For troubleshooting ipaddress objects uncomment lines below 
    # print(dir(ffa_supernet))
    # print(ffa_supernet.netmask)
    # print(ffa_supernet.network_address)
    # print(ffa_supernet.with_prefixlen)
    # print(ffa_supernet.num_addresses)
    # print(ffa_supernet.prefixlen)
    
    idmz_subnets = []
    transit_subnets = []
    ot_subnets = []

    networks = {}
    networks['idmz_subnets'] = idmz_subnets
    networks['transit_subnets'] = transit_subnets
    networks['ot_subnets'] = ot_subnets

    print(Fore.BLUE + Style.BRIGHT)
    print(Fore.RED + Style.BRIGHT + f"FFA Block Found in Parameters File: {summary_data['ffa_block_xls']}\n" + Fore.RESET)

    if ffa_supernet.prefixlen == 21:
        
        # splits the supernet into a list of /24's. calling [0] gets the first /24, [1] the second etc
        ffa_block_24s = list(ffa_supernet.subnets(new_prefix=24))
        print(f"\n{len(ffa_block_24s)} /24's in the /{ffa_supernet.prefixlen} ffa block")
        
        # splits the first /24 into /28's and returns the first 12 /28's
        idmz_subnets = list(ffa_block_24s[0].subnets(new_prefix=28))[:6]
        row = 16
        for subnet in idmz_subnets:
            row += 1
            
            summary_sheet[f'C{row}'] = subnet.with_prefixlen  # Network 
            summary_sheet[f'D{row}'] = str(subnet[1])  # Gateway
            summary_sheet[f'E{row}'] = str(subnet[2])  # First IP After Gateway
            summary_sheet[f'F{row}'] = str(subnet[-2])  # Last IP 
            summary_sheet[f'G{row}'] = str(subnet.netmask)  
       
        # Printing Stuff for the iDMZ Subnets
        print(Fore.BLUE + Style.BRIGHT + f"The first {len(idmz_subnets)} will be used for idmz subnets")
        i = 0
        for subnet in idmz_subnets:
            # Pulls the network address out of the network object and prints it
            i += 1
            print(Fore.GREEN + Style.BRIGHT + f"idmz subnet {i} " + subnet.with_prefixlen + Fore.RESET)
        
        # splits the first /24 into /29's and returns the last 4 /29's
        transit_subnets = list(ffa_block_24s[0].subnets(new_prefix=29))[-4:]
        
        row = 10
        for subnet in transit_subnets:
            row += 1

            summary_sheet[f'C{row}'] = subnet.with_prefixlen  # Network 
            summary_sheet[f'D{row}'] = 'N/A'  # Gateway
            summary_sheet[f'E{row}'] = str(subnet[2])  # First IP After Gateway
            summary_sheet[f'F{row}'] = str(subnet[-2])  # Last IP 
            summary_sheet[f'G{row}'] = str(subnet.netmask)  
        
        ot_management = list(ffa_block_24s[1].subnets(new_prefix=25))[0]
        
        summary_sheet['C24'] = ot_management.with_prefixlen  # Network
        summary_sheet['D24'] = str(ot_management[1])  # Gateway 
        summary_sheet['E24'] = str(ot_management[2])  # First IP After Gateway
        summary_sheet['F24'] = str(ot_management[-2])  # Last IP 
        summary_sheet['G24'] = str(ot_management.netmask)  
        
        ot_virtual_management = list(ffa_block_24s[1].subnets(new_prefix=25))[1]
        
        summary_sheet['C25'] = ot_virtual_management.with_prefixlen  # Network
        summary_sheet['D25'] = str(ot_virtual_management[1])  # Gateway 
        summary_sheet['E25'] = str(ot_virtual_management[2])  # First IP After Gateway
        summary_sheet['F25'] = str(ot_virtual_management[-2])  # Last IP 
        summary_sheet['G25'] = str(ot_virtual_management.netmask)  

        ot_vmotion = list(ffa_block_24s[2].subnets(new_prefix=25))[0]
        ot_vsan = list(ffa_block_24s[2].subnets(new_prefix=25))[1]
        
        ot_servers_1 = ot_subnets.append(list(ffa_block_24s[3].subnets(new_prefix=25))[0])
        ot_servers_2 = ot_subnets.append(list(ffa_block_24s[3].subnets(new_prefix=25))[1])
        ot_subnet_1 = ot_subnets.append(list(ffa_block_24s[4].subnets(new_prefix=25))[0])
        ot_subnet_2 = ot_subnets.append(list(ffa_block_24s[4].subnets(new_prefix=25))[1])
        ot_subnet_3 = ot_subnets.append(list(ffa_block_24s[5].subnets(new_prefix=25))[0])
        ot_subnet_4 = ot_subnets.append(list(ffa_block_24s[5].subnets(new_prefix=25))[1])
        ot_subnet_5 = ot_subnets.append(list(ffa_block_24s[6].subnets(new_prefix=25))[0])
        ot_subnet_6 = ot_subnets.append(list(ffa_block_24s[6].subnets(new_prefix=25))[1])
        ot_subnet_7 = ot_subnets.append(list(ffa_block_24s[7].subnets(new_prefix=25))[0])
        ot_subnet_8 = ot_subnets.append(list(ffa_block_24s[7].subnets(new_prefix=25))[1])
        # ot_subnet_9 = ot_subnets.append(list(ffa_block_24s[8].subnets(new_prefix=25))[0])
        # ot_subnet_10 = ot_subnets.append(list(ffa_block_24s[8].subnets(new_prefix=25))[1])

        row = 28
        for subnet in ot_subnets:
            row += 1                
            
            summary_sheet[f'C{row}'] = subnet.with_prefixlen  # Network
            summary_sheet[f'D{row}'] = str(subnet[1])  # Gateway 
            summary_sheet[f'E{row}'] = str(subnet[2])  # First IP After Gateway
            summary_sheet[f'F{row}'] = str(subnet[-2])  # Last IP 
            summary_sheet[f'G{row}'] = str(subnet.netmask)  

        # Subnets 9 & 10 are N/A for a /21
        for row in [39, 40]:
            
            summary_sheet[f'C{row}'] = 'N/A'  # Network 
            summary_sheet[f'D{row}'] = 'N/A'  # Gateway
            summary_sheet[f'E{row}'] = 'N/A'  # First IP After Gateway
            summary_sheet[f'F{row}'] = 'N/A'  # Last IP 
            summary_sheet[f'G{row}'] = 'N/A'  

        # Printing Stuff for the Transit Subnets
        print(Fore.BLUE + Style.BRIGHT + f"The Last 4 {len(transit_subnets)} will be used for transit subnets")
        i = 0
        for subnet in transit_subnets:
            # Pulls the network address out of the network object and prints it
            i += 1
            print(Fore.GREEN + f"transit subnet {i} " + subnet.with_prefixlen + Fore.RESET)
    
    elif ffa_supernet.prefixlen <= 20:
        
        # splits the supernet into a list of /24's. calling [0] gets the first /24, [1] the second etc
        ffa_block_24s = list(ffa_supernet.subnets(new_prefix=24))
        
        print(Fore.BLUE + Style.BRIGHT + f"{len(ffa_block_24s)} /24's in the /{ffa_supernet.prefixlen} ffa block")
        
        # splits the first /24 into /28's and returns the first 6 as that's the max iDMZ we ever use
        idmz_subnets = list(ffa_block_24s[0].subnets(new_prefix=28))[:6]
        
        # splits the second /24 into /28's and returns the first 12 /28's then extends (adds) the idmz_subnets list < Not in Use 6 iDMZ Max
        # idmz_subnets.extend(list(ffa_block_24s[1].subnets(new_prefix=28))[:-4])

        row = 16
        for subnet in idmz_subnets:
            row += 1
            summary_sheet[f'C{row}'] = subnet.with_prefixlen  # Network 
            summary_sheet[f'D{row}'] = str(subnet[1])  # Gateway
            summary_sheet[f'E{row}'] = str(subnet[2])  # First IP After Gateway
            summary_sheet[f'F{row}'] = str(subnet[-2])  # Last IP 
            summary_sheet[f'G{row}'] = str(subnet.netmask)  
        
        # Printing Stuff for the iDMZ Subnets
        print(Fore.BLUE + Style.BRIGHT + f"{len(idmz_subnets)} will be used for idmz subnets")
        i = 0
        for subnet in idmz_subnets:
            # Pulls the network address out of the network object and prints it
            i += 1
            print(Fore.GREEN + Style.BRIGHT + f"idmz subnet {i} " + subnet.with_prefixlen + Fore.RESET)
        
        # splits the second /24 into /28's and returns the last 4 /28's
        transit_subnets = list(ffa_block_24s[1].subnets(new_prefix=29))[-4:]
        
        row = 10
        for subnet in transit_subnets:
            row += 1
            summary_sheet[f'C{row}'] = subnet.with_prefixlen  # Network 
            summary_sheet[f'D{row}'] = 'N/A'  # Gateway
            summary_sheet[f'E{row}'] = str(subnet[2])  # First IP After Gateway
            summary_sheet[f'F{row}'] = str(subnet[-2])  # Last IP 
            summary_sheet[f'G{row}'] = str(subnet.netmask)  
        
        ot_management = list(ffa_block_24s[1].subnets(new_prefix=25))[0]
        summary_sheet['C24'] = ot_management.with_prefixlen  # Network 
        summary_sheet['D24'] = str(ot_management[1])  # Gateway
        summary_sheet['E24'] = str(ot_management[2])  # First IP After Gateway
        summary_sheet['F24'] = str(ot_management[-2])  # Last IP 
        summary_sheet['G24'] = str(ot_management.netmask)  
        
        ot_virtual_management = list(ffa_block_24s[1].subnets(new_prefix=25))[1]
        summary_sheet['C25'] = ot_virtual_management.with_prefixlen  # Network 
        summary_sheet['D25'] = str(ot_virtual_management[1])  # Gateway
        summary_sheet['E25'] = str(ot_virtual_management[2])  # First IP After Gateway
        summary_sheet['F25'] = str(ot_virtual_management[-2])  # Last IP 
        summary_sheet['G25'] = str(ot_virtual_management.netmask)  

        ot_vmotion = list(ffa_block_24s[2].subnets(new_prefix=25))[0]
        ot_vsan = list(ffa_block_24s[2].subnets(new_prefix=25))[1]
        
        ot_servers_1 = ot_subnets.append(list(ffa_block_24s[3].subnets(new_prefix=25))[0])
        ot_servers_2 = ot_subnets.append(list(ffa_block_24s[3].subnets(new_prefix=25))[1])
        ot_subnet_1 = ot_subnets.append(list(ffa_block_24s[4].subnets(new_prefix=25))[0])
        ot_subnet_2 = ot_subnets.append(list(ffa_block_24s[4].subnets(new_prefix=25))[1])
        ot_subnet_3 = ot_subnets.append(list(ffa_block_24s[5].subnets(new_prefix=25))[0])
        ot_subnet_4 = ot_subnets.append(list(ffa_block_24s[5].subnets(new_prefix=25))[1])
        ot_subnet_5 = ot_subnets.append(list(ffa_block_24s[6].subnets(new_prefix=25))[0])
        ot_subnet_6 = ot_subnets.append(list(ffa_block_24s[6].subnets(new_prefix=25))[1])
        ot_subnet_7 = ot_subnets.append(list(ffa_block_24s[7].subnets(new_prefix=25))[0])
        ot_subnet_8 = ot_subnets.append(list(ffa_block_24s[7].subnets(new_prefix=25))[1])
        ot_subnet_9 = ot_subnets.append(list(ffa_block_24s[8].subnets(new_prefix=25))[0])
        ot_subnet_10 = ot_subnets.append(list(ffa_block_24s[8].subnets(new_prefix=25))[1])

        row = 28
        for subnet in ot_subnets:
            row += 1                
            summary_sheet[f'C{row}'] = subnet.with_prefixlen  # Network 
            summary_sheet[f'D{row}'] = str(subnet[1])  # Gateway
            summary_sheet[f'E{row}'] = str(subnet[2])  # First IP After Gateway
            summary_sheet[f'F{row}'] = str(subnet[-2])  # Last IP 
            summary_sheet[f'G{row}'] = str(subnet.netmask)  

        # Printing Stuff for the Transit Subnets
        print(Fore.BLUE + Style.BRIGHT + f"{len(transit_subnets)} will be used for transit subnets")
        i = 0
        for subnet in transit_subnets:
            # Pulls the network address out of the network object and prints it
            i += 1
            print(Fore.GREEN + f"transit subnet {i} " + subnet.with_prefixlen + Fore.RESET)
            
    else:
        print(Fore.RED + Style.BRIGHT + f"The prefix Length /{ffa_supernet.prefixlen} is not valid" + Fore.RESET)
        exit()
    
    r1 = input(f"\nDo you want to Write these networks to '{filename}'? This will overwrite any existing values. [y/n][default n]: ").lower()
    if r1 == 'y':
        workbook.save(filename)
        # workbook.save(f"{filename[:-5]}_{summary_data['city'][:6]}.xlsx")
        workbook.close()
        r2 = input(f"Values saved to {filename}. Are modifications of networks needed before continuing? [y/n][default n]: ").lower()
        if r2 == 'y':
            print(Fore.RED + f"\nPlease make modifications to {filename} and restart, exiting!" + Fore.RESET)
            exit()
        print(Fore.RED + f"\nPopulating {filename} with IP Info\n" + Fore.RESET)
        generate_ip_address_info(summary_data)
    return filename


def generate_ip_address_info(summary_data):
    """This will use the Networks found on the Summary sheet to Populate IP Info and VLAN sheets.
        
    Args:
        summary_data ([dict]): info from summary sheet
        VLAN and Description Required plus (Subnet Mask + 1 or more of [starting, ending, gateway ip]) or (Network in CIDR Format)
    """    
 
    filename = summary_data['filename']
    state = summary_data['site']
    city = summary_data['city']
    ffa_block_xls = summary_data['ffa_block_xls']

    column_names = []
    for cell in summary_sheet[10]:
        column_names.append(cell.value)
    # print(column_names)
    
    # These column names must be in row two of Summary Sheet
    vlan_index = column_names.index("VLAN") + 1
    description_index = column_names.index("Description") + 1
    gateway_index = column_names.index("Gateway Address") + 1
    network_index = column_names.index("Network") + 1
    starting_index = column_names.index("Starting IP Address") + 1
    ending_index = column_names.index("Ending IP Address") + 1
    mask_index = column_names.index("Subnet Mask") + 1
    
    # summary_sheet['B7'] = f"pb{city}{state}".lower()

    ffa_supernet = ipaddress.ip_network(ffa_block_xls)
    
    idmz_subnets = []
    transit_subnets = []
    ot_subnets = []

    networks = {}
    networks['transit_subnets'] = transit_subnets
    networks['idmz_subnets'] = idmz_subnets
    networks['ot_subnets'] = ot_subnets
    
    # Update Transit subnets
    row_num = 11
    for row in summary_sheet.iter_rows(values_only=True, min_row=11, max_row=16):

        # Skip rows that are don't have enough information to create a network object
        if row[vlan_index] is None or row[description_index] is None or row[network_index] == 'N/A':
            continue
        
        if row[mask_index] is None and row[network_index] is None:
            continue

        if row[mask_index] == "N/A" and row[network_index] is None:
            continue
        
        subnet = {}
        subnet['description'] = row[description_index]
        subnet['vlan'] = row[vlan_index]
        network = None
        
        if row[network_index] is not None and row[network_index] != "xxx.xxx.xxx.xxx/xx":
            network = ipaddress.ip_network(row[network_index])
        elif row[mask_index]:
            if row[ending_index]:
                network = ipaddress.ip_network(f"{row[ending_index]}/{row[mask_index]}", strict=False)
            elif row[starting_index]:
                network = ipaddress.ip_network(f"{row[starting_index]}/{row[mask_index]}", strict=False)
            elif row[gateway_index]:
                network = ipaddress.ip_network(f"{row[gateway_index]}/{row[mask_index]}", strict=False)
        
        if network is None:
            print(f"Could not determine network information for vlan {row[vlan_index]} please resolve.")
            continue
        
        # Save Subnet information for future use
        subnet['network'] = network
        subnet['network_w_prefix'] = network.with_prefixlen  # Network 
        subnet['netmask'] = str(network.netmask)  # Subnet Mask
        subnet['gateway'] = str(network[1])  # Gateway
        subnet['starting_ip'] = str(network[2])  # First IP After Gateway
        subnet['ending_ip'] = str(network[-2])  # Last IP 
        transit_subnets.append(subnet)

        # Write missing information to excel 
        summary_sheet.cell(row=row_num, column=gateway_index, value='N/A')  # Gateway
        summary_sheet.cell(row=row_num, column=network_index, value=subnet['network_w_prefix'])  # Network
        summary_sheet.cell(row=row_num, column=starting_index, value=subnet['starting_ip'])  # First IP After Gateway
        summary_sheet.cell(row=row_num, column=ending_index, value=subnet['ending_ip'])  # Last IP 
        summary_sheet.cell(row=row_num, column=mask_index, value=subnet['netmask'])  # Subnet Mask

        row_num += 1

    # Update IDMZ subnets
    row_num = 17
    for row in summary_sheet.iter_rows(values_only=True, min_row=17, max_row=22):
        # if row[0] is None or row[3] == 'N/A' or row[3] is None:
        if row[vlan_index] is None or row[network_index] == 'N/A' or row[description_index] is None:
            continue

        if row[mask_index] is None and row[network_index] is None:
            continue

        if row[mask_index] == "N/A" and row[network_index] is None:
            continue

        subnet = {}
        subnet['description'] = row[description_index]
        subnet['vlan'] = row[vlan_index]
        network = None
        
        if row[network_index] is not None and row[network_index] != "xxx.xxx.xxx.xxx/xx":
            network = ipaddress.ip_network(row[network_index])
        elif row[mask_index]:
            if row[ending_index]:
                network = ipaddress.ip_network(f"{row[ending_index]}/{row[mask_index]}", strict=False)
            elif row[starting_index]:
                network = ipaddress.ip_network(f"{row[starting_index]}/{row[mask_index]}", strict=False)
            elif row[gateway_index]:
                network = ipaddress.ip_network(f"{row[gateway_index]}/{row[mask_index]}", strict=False)
        
        if network is None:
            print(f"Could not determine network information for vlan {row[vlan_index]} please resolve.")
            continue
        
        # Save Subnet information for future use
        subnet['network'] = network
        subnet['network_w_prefix'] = network.with_prefixlen  # Network 
        subnet['netmask'] = str(network.netmask)  # Subnet Mask
        subnet['gateway'] = str(network[1])  # Gateway
        subnet['starting_ip'] = str(network[2])  # First IP After Gateway
        subnet['ending_ip'] = str(network[-2])  # Last IP 
        idmz_subnets.append(subnet)
        
        # Write missing information to excel 
        summary_sheet.cell(row=row_num, column=gateway_index, value='N/A')  # Gateway
        summary_sheet.cell(row=row_num, column=network_index, value=subnet['network_w_prefix'])  # Network
        summary_sheet.cell(row=row_num, column=starting_index, value=subnet['starting_ip'])  # First IP After Gateway
        summary_sheet.cell(row=row_num, column=ending_index, value=subnet['ending_ip'])  # Last IP 
        summary_sheet.cell(row=row_num, column=mask_index, value=subnet['netmask'])  # Subnet Mask

        row_num += 1

    # Update OT subnets

    for row in summary_sheet.iter_rows(min_row=24, max_row=50):
        # if row[0] is None or row[3] == 'N/A' or row[3] is None or row[network_index] == "xxx.xxx.xxx.xxx/xx":
        if row[vlan_index].value is None or row[description_index].value is None or row[network_index].value == "N/A":
            continue

        if row[mask_index].value is None and row[network_index].value is None:
            continue
        
        if row[mask_index].value == "N/A" and row[network_index].value is None:
            continue

        subnet = {}
        subnet['description'] = row[description_index].value
        subnet['vlan'] = row[vlan_index].value
        network = None
        
        if row[network_index].value is not None and row[network_index].value != "xxx.xxx.xxx.xxx/xx":
            network = ipaddress.ip_network(row[network_index].value)
        elif row[mask_index]:
            if row[ending_index].value:
                network = ipaddress.ip_network(f"{row[ending_index].value}/{row[mask_index].value}", strict=False)
            elif row[starting_index].value:
                network = ipaddress.ip_network(f"{row[starting_index].value}/{row[mask_index].value}", strict=False)
            elif row[gateway_index].value:
                network = ipaddress.ip_network(f"{row[gateway_index].value}/{row[mask_index].value}", strict=False)
        
        if network is None:
            print(f"Could not determine network information for vlan {row[vlan_index].value} please resolve.")
            continue
        
        # Save Subnet information for future use
        subnet['network'] = network
        subnet['network_w_prefix'] = network.with_prefixlen  # Network
        subnet['netmask'] = str(network.netmask)  # Subnet Mask
        subnet['gateway'] = str(network[1])  # Gateway
        subnet['starting_ip'] = str(network[2])  # First IP After Gateway
        subnet['ending_ip'] = str(network[-2])  # Last IP
        ot_subnets.append(subnet)

        # Write missing information to excel 
        summary_sheet.cell(row=row_num, column=gateway_index, value='N/A')  # Gateway
        summary_sheet.cell(row=row_num, column=network_index, value=subnet['network_w_prefix'])  # Network
        summary_sheet.cell(row=row_num, column=starting_index, value=subnet['starting_ip'])  # First IP After Gateway
        summary_sheet.cell(row=row_num, column=ending_index, value=subnet['ending_ip'])  # Last IP 
        summary_sheet.cell(row=row_num, column=mask_index, value=subnet['netmask'])  # Subnet Mask


    # Update BT Connections Sheet
    fw01_name = summary_data['fw01']
    fw02_name = summary_data['fw02']
    
    for i in range(4, 10):
        bt_connection_sheet[f"A{i}"] = fw01_name
    for i in range(10, 16):
        bt_connection_sheet[f"A{i}"] = fw02_name   

    bt_connection_sheet['F4'] = f"{fw01_name[4:]}cs01"
    bt_connection_sheet['F5'] = f"pb{summary_data['city'][:6]}{summary_data['state']}38r01-ot".lower()
    bt_connection_sheet['F6'] = f"{fw01_name[4:]}dmz01"
    bt_connection_sheet['F7'] = f"{fw01_name[4:]}as0x"
    bt_connection_sheet['F8'] = fw02_name
    bt_connection_sheet['F9'] = f"{fw01_name[4:]}as0x"
    bt_connection_sheet['F10'] = f"{fw01_name[4:]}cs01"
    bt_connection_sheet['F11'] = f"pb{summary_data['city'][:6]}{summary_data['state']}38r01-ot".lower()
    bt_connection_sheet['F12'] = f"{fw01_name[4:]}dmz01"
    bt_connection_sheet['F13'] = f"{fw01_name[4:]}as0x"
    bt_connection_sheet['F14'] = fw01_name
    bt_connection_sheet['F15'] = f"{fw01_name[4:]}as0x"
    bt_connection_sheet['D4'] = f"{'.'.join(summary_sheet['E11'].value.split('.')[:3])}.237/29"
    bt_connection_sheet['D5'] = f"{'.'.join(summary_sheet['E12'].value.split('.')[:3])}.209/29"
    bt_connection_sheet['D6'] = f"{'.'.join(summary_sheet['E17'].value.split('.')[:3])}.2/28"
    bt_connection_sheet['D10'] = f"{'.'.join(summary_sheet['E11'].value.split('.')[:3])}.238/29"
    bt_connection_sheet['D11'] = f"{'.'.join(summary_sheet['E12'].value.split('.')[:3])}.210/29"
    bt_connection_sheet['D12'] = f"{'.'.join(summary_sheet['E17'].value.split('.')[:3])}.3/28"
    bt_connection_sheet['E19'] = f"ip route {ffa_supernet.network_address} {ffa_supernet.netmask} {bt_connection_sheet['D4'].value[:-3]}"
    bt_connection_sheet['E21'] = f"route corplan 0.0.0.0 0.0.0.0 {'.'.join(str(ffa_supernet.network_address).split('.')[:3])}.233"
    bt_connection_sheet['E22'] = f"route mfglan {ffa_supernet.network_address} {ffa_supernet.netmask} {'.'.join(str(ffa_supernet.network_address).split('.')[:3])}.212"
    bt_connection_sheet['E23'] = f"ip route 0.0.0.0 0.0.0.0 {summary_sheet['F12'].value}"
        
    # Update IDMZ Vlans Sheet
    row = 4
    vlan_id = 653
    for network in idmz_subnets:
        subnet = ipaddress.ip_network(network['network'])
        vlan_id += 1
        for host in subnet.hosts():
            row += 1
            # if str(subnet[1]) == str(host):
            #     idmz_vlans_sheet[f'B{row}'] = 'Gateway Address'
            # elif str(subnet[-2]) == str(host):
            #     idmz_vlans_sheet[f'B{row}'] = 'Last IP Address'
            idmz_vlans_sheet[f'A{row}'] = str(host)
            idmz_vlans_sheet[f'D{row}'] = vlan_id
    
    # Update Transit Vlans Sheet
    row = 4
    vlan_ids = [65, 60]
    index = 0
    for network in transit_subnets[:2]:
        # print(subnet.with_prefixlen)
        subnet = ipaddress.ip_network(network['network'])
        for host in subnet.hosts():
            row += 1
            # print(str(host))
            transit_vlans_sheet[f'A{row}'] = str(host)
            transit_vlans_sheet[f'D{row}'] = vlan_ids[index]
        index += 1
    transit_vlans_sheet['A18'] = str(list(ipaddress.ip_network(transit_subnets[-1]['network']).hosts())[1])
    transit_vlans_sheet['A19'] = str(list(ipaddress.ip_network(transit_subnets[-1]['network']).hosts())[2])

    networks = idmz_subnets + ot_subnets
    # for row in summary_sheet.iter_rows(max_col=11, min_row=24):
    #     if row[0].value and row[3].value and row[3].value != 'N/A':
    #         vlans.append({'id': row[0].value, 'description': row[1].value, 'network': row[3].value, 'gateway': row[4].value})
            
    for network in networks:
        # print(network['vlan'])
        if network['network'] is None or network['network'] == 'N/A' or 'iDMZ' in network['description']:
            continue
        sheet_found = False
        # print(network['vlan'])
        for sheetname in workbook.sheetnames:
            if sheetname == 'VLAN - 600 Management':
                # print('600 found')
                sheet = workbook[sheetname]
                sheet["B5"] = f"pb{summary_data['city'][:6]}{summary_data['state']}38r01-ot".upper()
                sheet["B15"] = f"pb{summary_data['city'][:6]}{summary_data['state']}38s11-ot".upper()
                sheet["B16"] = f"pb{summary_data['city'][:6]}{summary_data['state']}36s12-ot".upper()
                sheet["B17"] = f"pb{summary_data['city'][:6]}{summary_data['state']}36s13-ot".upper()
            
            elif sheetname == 'VLAN 601 OT Virtual Management':
                # print('601 found')
                sheet = workbook[sheetname]
                sheet["B5"] = f"PF{summary_data['bt_code'].replace('-', '')}38r01-ot".upper()
                
                sheet["B8"] = f"PF{summary_data['bt_code'].replace('-', '')}IPDU01".upper()
                sheet["B9"] = f"PF{summary_data['bt_code'].replace('-', '')}IPDU02".upper()
                sheet["B10"] = f"PF{summary_data['bt_code'].replace('-', '')}IUPS01".upper()

                sheet["B14"] = f"PF{summary_data['bt_code'].replace('-', '')}VIM01".upper()
                sheet["B15"] = f"PF{summary_data['bt_code'].replace('-', '')}IMH01".upper()
                sheet["B16"] = f"PF{summary_data['bt_code'].replace('-', '')}IMH02".upper()
                sheet["B17"] = f"PF{summary_data['bt_code'].replace('-', '')}IMH03".upper()
                sheet["B18"] = f"PF{summary_data['bt_code'].replace('-', '')}IMH04".upper()

                sheet["B25"] = f"PF{summary_data['bt_code'].replace('-', '')}VMH01".upper()
                sheet["B26"] = f"PF{summary_data['bt_code'].replace('-', '')}VMH02".upper()
                sheet["B27"] = f"PF{summary_data['bt_code'].replace('-', '')}VMH03".upper()
                sheet["B28"] = f"PF{summary_data['bt_code'].replace('-', '')}VMH04".upper()
                sheet["B29"] = f"PF{summary_data['bt_code'].replace('-', '')}VMH05".upper()

                sheet["B37"] = f"PF{summary_data['bt_code'].replace('-', '')}WBAK01".upper()
                sheet["B38"] = f"PF{summary_data['bt_code'].replace('-', '')}WLIC01".upper()
                sheet["B39"] = f"PF{summary_data['bt_code'].replace('-', '')}WFIL01".upper()  
            
            elif sheetname == "VLAN 606 - OT Servers 1":
                # print('606 found')
                sheet = workbook[sheetname]
                sheet["B5"] = f"pb{summary_data['city'][:6]}{summary_data['state']}38r01-ot".upper()
                
                sheet["B15"] = f"pf{summary_data['city'][:6]}{summary_data['state']}WAD01".upper()
                sheet["B16"] = f"pf{summary_data['city'][:6]}{summary_data['state']}WAD02".upper()
                sheet["B17"] = f"pf{summary_data['city'][:6]}{summary_data['state']}WSQL01".upper()
                sheet["B18"] = f"pf{summary_data['city'][:6]}{summary_data['state']}WZPI01".upper()
                
                sheet["B27"] = f"pf{summary_data['city'][:6]}{summary_data['state']}FTD01".upper()
                sheet["B28"] = f"pf{summary_data['city'][:6]}{summary_data['state']}FTAC01".upper()
                sheet["B29"] = f"pf{summary_data['city'][:6]}{summary_data['state']}FTAG01".upper()
                sheet["B30"] = f"pf{summary_data['city'][:6]}{summary_data['state']}FTAG02".upper()

                sheet["B34"] = f"pf{summary_data['city'][:6]}{summary_data['state']}FHMI01".upper()
                sheet["B35"] = f"pf{summary_data['city'][:6]}{summary_data['state']}FHMI02".upper()
                sheet["B36"] = f"pf{summary_data['city'][:6]}{summary_data['state']}FHMI03".upper()
                sheet["B37"] = f"pf{summary_data['city'][:6]}{summary_data['state']}FHMI04".upper()
                sheet["B38"] = f"pf{summary_data['city'][:6]}{summary_data['state']}FHMI05".upper()

                sheet["B44"] = f"pf{summary_data['city'][:6]}{summary_data['state']}WTM01".upper()
                
                sheet["B49"] = f"pf{summary_data['city'][:6]}{summary_data['state']}WRDS01".upper()
                sheet["B50"] = f"pf{summary_data['city'][:6]}{summary_data['state']}WRDS02".upper()

                sheet["B54"] = f"pf{summary_data['city'][:6]}{summary_data['state']}FBCH01".upper()
                sheet["B55"] = f"pf{summary_data['city'][:6]}{summary_data['state']}FHIS01".upper()
                sheet["B56"] = f"pf{summary_data['city'][:6]}{summary_data['state']}WRLE01".upper()

                sheet["B76"] = f"pf{summary_data['city'][:6]}{summary_data['state']}WVSE01".upper()
                sheet["B77"] = f"pf{summary_data['city'][:6]}{summary_data['state']}WSIE01".upper()
                sheet["B78"] = f"pf{summary_data['city'][:6]}{summary_data['state']}WEWS01".upper()
                sheet["B79"] = f"pf{summary_data['city'][:6]}{summary_data['state']}WEWS02".upper()
                sheet["B80"] = f"pf{summary_data['city'][:6]}{summary_data['state']}WVSE01".upper()
                sheet["B81"] = f"pf{summary_data['city'][:6]}{summary_data['state']}WSIE01".upper()
                sheet["B82"] = f"pf{summary_data['city'][:6]}{summary_data['state']}WEWS01".upper()
                sheet["B83"] = f"pf{summary_data['city'][:6]}{summary_data['state']}WEWS02".upper()
           
            if str(network['vlan']) in sheetname:
                # print(f"{vlan['id']} is in {sheetname}")
                sheet_found = True
                sheet = workbook[sheetname]

                sheet['B1'] = f"{network['description']} / VLAN {network['vlan']} / {network['network']}"
                sheet['B2'] = network['gateway']
                sheet['B3'] = network['netmask']
                
                row = 4

                print(Fore.BLUE + f"Writing ip's for VLAN {network['vlan']} - {network['description']} {network['network']}" + Fore.RESET)
                sheet.title = f"VLAN {network['vlan']} {network['description']}"
                # print(f"{ipaddress.ip_network(network['network']).num_addresses} usable addresses" + Fore.RESET)
                # print(network['network'])
                for host in ipaddress.ip_network(network['network']).hosts():
                    row += 1
                    sheet[f'A{row}'] = str(host)
            
        if sheet_found is False:
            print(Fore.RED + f"\nVLAN {network['vlan']} - {network['description']} {network['network']} was not found in any sheet, creating" + Fore.RESET)
            template_sheet = workbook['VLAN Template']
            sheet = workbook.copy_worksheet(template_sheet)
            sheet['B1'] = f"{network['description']} / VLAN {network['vlan']} / {network['network']}"
            sheet['B2'] = network['gateway']
            sheet['B3'] = network['netmask']
            row = 4

            print(Fore.BLUE + f"Writing ip's for VLAN {network['vlan']} - {network['description']} {network['network']}" + Fore.RESET)
            sheet.title = f"VLAN {network['vlan']} {network['description']}"
            # print(f"{ipaddress.ip_network(network['network']).num_addresses} usable addresses" + Fore.RESET)
            # print(network['network'])
            for host in ipaddress.ip_network(network['network']).hosts():
                row += 1
                sheet[f'A{row}'] = str(host)
    

    # mod_filename = f"{filename[:-5]}_{summary_data['city'][:6]}.xlsx"
    # print(Fore.BLUE + f"\nsaving {mod_filename}!" + Fore.RESET)
    # workbook.save(mod_filename)
    workbook.save(filename)
    workbook.close()


def generate_portmap_data(hosts, summary_data):
    """ Gets info from config files and arp files to create portmap data

    Args:
        hosts ([list]): list of nornir host objects
        summary_data ([type]): all of the data on the summary page 
    """    

    print(Fore.BLUE + "Starting Portmap" + Fore.RESET)
    
    # load the excel document 
    filename = summary_data['filename']
    workbook = load_workbook(filename=filename)
    
    portmap_template_sheet = workbook["Port Map Template"]
    column_names = []
    for cell in portmap_template_sheet[1]:
        column_names.append(cell.value)
    # print(column_names)

    # These column names must be in row two of Port Map Template Sheet
    device_name_index = column_names.index("Name")
    device_ip_index = column_names.index("IP Address")
    
    interface_name_index = column_names.index("Interface")
    interface_vlan_index = column_names.index("VLAN")
    interface_status_index = column_names.index("Status")
    interface_config_index = column_names.index("Config Lines")
    interface_neighbor_index = column_names.index("CDP Neighbor")
    interface_description_index = column_names.index("Interface Description")

    visio_portmap_index = column_names.index("Visio Portmap")
    connected_device_description_index = column_names.index("Connected Device Description")
    connected_device_remote_interface_index = column_names.index("Remote Interface")
       
    visio_sheet = workbook["Visio Data"]
    column_names = []
    for cell in visio_sheet[1]:
        column_names.append(cell.value)
    # print(visio_sheet)
    
    # These column names must be in row two of Physical Connections visio data sheet    
    page_index = column_names.index("Page")
    description_index = column_names.index("Description:")
    name_index = column_names.index("Name:")
    location_index = column_names.index("Location:")
    model_index = column_names.index("Model:")
    ip_index = column_names.index("IP Address:")
    mask_index = column_names.index("Subnet Mask:")
    vlan_index = column_names.index("VLAN:")
    data_index = column_names.index("Data")  
      
    vlan_template_sheet = workbook["VLAN Template"]
    tcolumn_names = []
    
    for cell in vlan_template_sheet[4]:
        tcolumn_names.append(cell.value)

    # These column names must be in row 4 of vlan Template Sheet data sheet 
    t_mac_index = tcolumn_names.index("MAC Address") + 1
    t_switch_index = tcolumn_names.index("Switch") + 1
    t_interface_index = tcolumn_names.index("Switchport") + 1
    t_interface_description_index = tcolumn_names.index("Interface Description") + 1

    # print(visio_sheet)
    # num_idfs = int(input(Fore.BLUE + "How Many IDF's are there? 0-8: "))
    # visio_switches_functions = ['FW01', 'FW02', 'DMZ01', 'MDF01', 'IDC01']
    # for i in range(num_idfs):
    #     visio_switches_functions.append(f"IDF0{i + 1}")
    # host_choices = hosts.copy()
    # for switch_func in visio_switches_functions:
    #     identified = False
    #     while identified is not True:
    #         print(Fore.RED + "\nSwitches in inventory:")
    #         print(Fore.BLUE + f"{list(host_choices.keys())}" + Fore.RESET)
    #         switch_name = input(f"From the list of switch names enter the name for {switch_func} [s] to skip: ") 
    #         if switch_name.lower() == 's':
    #             break
    #         elif switch_name in host_choices:
    #             print(Fore.RED + f"{switch_name} identified as {switch_func}" + Fore.RESET)
    #             hosts[switch_name]['function'] = switch_func
    #             identified = True
    #             del host_choices[switch_name]
    #         else:
    #             print(Fore.RED + f"{switch_name} not found in hosts" + Fore.RESET)
    
    directory = './Show Tech/MAC_IP'
    arps = {}
    
    # Get [IP - Mac - VLAN] Relationship for all known IP's From MAC_IP's

    print(f'{Fore.BLUE}Getting Known IPs from MAC_IP Files{Fore.RESET}')
    if not os.path.isdir(os.path.dirname(directory)):
        print(Fore.RED + f"{directory}  was not found!" + Fore.RESET)
    
    for root, dirs, files in os.walk(directory):
        for f in files:
            # print(f"root {root}, dirs {dirs}, file {f}")
            # print(f"found arp file {f}")
            with open(f"{root}/{f}", 'r') as a:
                arp_lines = a.readlines()
            for line in arp_lines:
                if line.startswith('Internet'):
                    # print(line)
                    ip = line.split()[1]
                    vlan = line.split()[-1].lstrip('Vlan')
                    arp_entry = re.search(r"(\w+)\s+(\d+\.\d+\.\d+\.\d+)\s+(\d+)\s+(\w+.\w+.\w+)\s+(\w+)\s+(\S+)", line,)
                    if arp_entry:

                        device_mac = arp_entry.group(4)
                        # print(device_mac)
                        if ip in arps.keys() and arps[ip]['mac'] == device_mac:
                            continue
                        arps[ip] = {}     
                        arps[ip]['mac'] = device_mac
                        arps[ip]['vlan'] = vlan 
    
    # Get [HOST - INT - IP - MAC - VLAN] Relationship from MAC_IP's
    hosts_missing_arp_files = hosts.copy()
    hostname = None
    for root, dirs, files in os.walk(directory):
        for f in files:

            with open(f"{root}/{f}", 'r') as a:
                arp_lines = a.readlines()
               
            for ip in arps.keys():
                for line in arp_lines:
                    
                    # set the active host if hostname found in line
                    for host in hosts:
                        if host in line:
                            hostname = host
                            # delete host from list for debugging
                            if host in hosts_missing_arp_files.keys():
                                del hosts_missing_arp_files[host]
                    if hostname is None:
                        continue

                    if arps[ip]['mac'] in line:

                        iface = line.split()[-1]
                        
                        # if iface.lower().startswith('p'):
                        #     continue
                        # elif iface.lower().startswith('v'):
                        #     continue
                        
                        if iface.lower() in hosts[hostname]['interfaces'].keys():
                            # print(f"{iface} found on {host}")
                            if 'connected_devices' not in hosts[hostname]['interfaces'][iface.lower()]:
                                hosts[hostname]['interfaces'][iface.lower()]['connected_devices'] = {}
                            hosts[hostname]['interfaces'][iface.lower()]['connected_devices'][ip] = {'mac': arps[ip]['mac'], 'vlan': arps[ip]['vlan']}
    
    for host in hosts_missing_arp_files:
        print(Fore.RED + f"Did not find a MAC_IP File for {host}" + Fore.RESET)    
    
    # Portmap - Per VLAN
    print(Fore.BLUE + "\nStarting Portmap - Per VLAN" + Fore.RESET)
    missing_ip_sheet = {}
    
    # loop through Hosts
    for host in hosts:
        num_connected = 0
        
        if 'interfaces' not in hosts[host]:
            print(f"no interfaces on {host}")
            continue
        
        for interface in hosts[host]['interfaces']:
            if 'connected_devices' not in hosts[host]['interfaces'][interface]:
                # print(f"No connected devices on {host} {interface}")
                continue
            
            connected_devices = hosts[host]['interfaces'][interface]['connected_devices']
            
            # print(len(connected_devices))
            if len(connected_devices) >= 1:
                for ip in connected_devices:
                    ip_vlan = connected_devices[ip]['vlan']
                    # print(ip_vlan)
                    
                    if ip_vlan == '65' or ip_vlan == '60':
                        continue
                    
                    num_connected += 1
                    sheet_found = False
                    ip_found = False
                    
                    for sheetname in workbook.sheetnames:
                        
                        if str(connected_devices[ip]['vlan']) in sheetname:

                            # print('vlan')
                            sheet_found = True
                            sheet = workbook[sheetname]
                            for row in sheet.iter_rows(min_row=5):
                                if ip == row[0].value:

                                    # print(f"{Fore.BLUE}{ip} found on {host} {interface} adding to {sheetname} sheet")
                                    ip_found = True
                                    sheet.cell(row=row[0].row, column=t_mac_index, value=connected_devices[ip]['mac'].upper())
                                    sheet.cell(row=row[0].row, column=t_interface_index, value=interface.capitalize())
                                    sheet.cell(row=row[0].row, column=t_switch_index, value=host.upper())

                                    if 'description' in hosts[host]['interfaces'][interface]:
                                        sheet.cell(row=row[0].row, column=t_interface_description_index, value=hosts[host]['interfaces'][interface]['description'])
                                
                    if sheet_found is False:
                        # print(f"No sheet found for {ip} on vlan {connected_devices[ip]['vlan']}")
                        pass
                    
                    if ip_found is False:
                        missing_ip_sheet[connected_devices[ip]['vlan']] = ip
             
        print(Fore.RED + f"{host} Has {num_connected} Connected Devices" + Fore.RESET)

    # if the ip found in data wasn't found on any vlans sheets
    print(Fore.RED + "\nStarting diagnostics for devices found in IP_MAC files that couldn't be added to spreadsheet." + Fore.RED)
    
    for vlan, ip in missing_ip_sheet.items():
        print(Fore.RED + Style.BRIGHT + f"\nIP {ip} (possibly others too) was Found in ARP Data on {Fore.BLUE} VLAN {vlan} {Fore.RED} but was not found on vlan sheet in {filename}!" + Fore.RESET)
        
        vlan_found = False
        address = None
        network = None
        
        for subnet in summary_data['all_subnets']:
            # print(Fore.BLUE + f"VLAN: {subnet['vlan']} Network {subnet['network_w_prefix']}" + Fore.RESET)
            # print(subnet['network_w_prefix'])
            
            if str(subnet['vlan']) == str(vlan):
                print(f"Found VLAN {vlan} on Summary Sheet")
                vlan_found = True
                network = ipaddress.IPv4Network(subnet['network_w_prefix'])
                address = ipaddress.ip_address(ip)
                if address and network:
                    
                    if address in network:
                        print(Fore.RED + f" The IP {ip} is a VALID address in VLAN {vlan} Network {subnet['network_w_prefix']} found on Summary Sheet\
                            \n Please Verify {vlan} ip sheet is created and contains IP {ip}!")
                    
                    else:
                        print(Fore.RED + f"The IP {ip} is an INVALID address in the VLAN {vlan} network {subnet['network_w_prefix']} found on Summary Sheet\
                            \n Verify VLAN {vlan} on Summary Sheet has the Correct Associated Network!\n")
            
        if vlan_found is False:
            print(Fore.RED + f"VLAN {vlan} was not found on Summary Sheet" + Fore.RESET)

        vlan_found = False
        address = None
        network = None
    # Portmap - Per VLAN - END ^

    # Portmap - Per Switch
    print(Fore.BLUE + "\nStarting Portmap - Per Switch" + Fore.RESET)
    visio_row = 150
    for host in hosts:
        sheet_found = False
        add_blank_line = False
        
        host_data = hosts[host]
        device_name = host
        device_ip = hosts[host].hostname
        portmap_lines = []
        portmap_visio_data = ""
        
        row = 2
        for sheetname in workbook.sheetnames:
            if host in sheetname:
                sheet_found = True
                portmap_sheet = workbook[sheetname] 
        
        if not sheet_found:
            portmap_sheet = workbook.copy_worksheet(portmap_template_sheet)
            portmap_sheet.title = f"{host} Port Map"

        
        if 'interfaces' in hosts[host].keys():
                       
            # loop for connected devices
            for iface_name, interface in host_data["interfaces"].items():   
                portmap_visio_data = ""
                
                interface_vlan = None
                interface_config = None
                interface_description = None
                
                interface_neighbor = None
                interface_neighbor_remote_interface = None
                
                connected_device_name = None
                                
                connected_device_description = None
                connected_device_remote_interface = None

                if 'connected_devices' in interface.keys():
                    
                    connected_devices = interface['connected_devices']

                    if len(connected_devices) >= 1:

                        portmap_visio_data += f"\n{iface_name.capitalize()}"
                        
                        for ip in connected_devices:
                            
                            portmap_visio_data += "\n"
                            portmap_visio_data += f"{ip}\n"
                            
                            if 'description' in interface.keys():
                                portmap_visio_data += f"{interface['description'].strip()}\n"
                            
                            portmap_visio_data += f"{connected_devices[ip]['mac'].upper()}\n"
                                  
                if 'config' in interface.keys():
                    interface_config = "\n".join(interface['config'])
                
                if 'description' in interface.keys():
                    interface_description = interface['description']

                if 'vlan' in interface.keys():
                    interface_vlan = interface['vlan']

                if 'neighbor' in interface.keys():
                    interface_neighbor = interface['neighbor']['remote_device']
                    connected_device_ip = interface['neighbor']['remote_address']
                    connected_device_remote_interface = interface['neighbor']['remote_iface']
                    
                    if not connected_device_name:
                        connected_device_name = interface_neighbor
                                                
                if add_blank_line:
                    # row += 1
                    portmap_sheet.insert_rows(row)
                    row += 1
                    add_blank_line = False

                portmap_sheet.insert_rows(row)
                portmap_sheet.cell(column=device_name_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                portmap_sheet.cell(column=device_name_index + 1, row=row, value=device_name)
                portmap_sheet.cell(column=device_ip_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                portmap_sheet.cell(column=device_ip_index + 1, row=row, value=device_ip)
                
                if iface_name:
                    portmap_sheet.cell(column=interface_name_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                    portmap_sheet.cell(column=interface_name_index + 1, row=row, value=iface_name.capitalize())
                
                if interface_vlan:
                    portmap_sheet.cell(column=interface_vlan_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                    portmap_sheet.cell(column=interface_vlan_index + 1, row=row, value=interface_vlan)
                
                if interface_config:
                    portmap_sheet.cell(column=interface_config_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                    portmap_sheet.cell(column=interface_config_index + 1, row=row, value=interface_config.strip())
                
                portmap_sheet.cell(column=interface_neighbor_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                portmap_sheet.cell(column=interface_neighbor_index + 1, row=row, value=interface_neighbor)
                                
                portmap_sheet.cell(column=interface_description_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                portmap_sheet.cell(column=interface_description_index + 1, row=row, value=interface_description)
                
                if portmap_visio_data:
                    portmap_sheet.cell(column=visio_portmap_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                    portmap_sheet.cell(column=visio_portmap_index + 1, row=row, value=portmap_visio_data.strip())
                    
                portmap_sheet.cell(column=connected_device_description_index + 1, row=row, value=connected_device_description)
                if connected_device_remote_interface:
                    portmap_sheet.cell(column=connected_device_remote_interface_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                    portmap_sheet.cell(column=connected_device_remote_interface_index + 1, row=row, value=connected_device_remote_interface.capitalize())
                
                portmap_sheet.row_dimensions[row].height = 15
                row += 1
            
            num = 1
            if 'vlans' in hosts[host].keys():
                for vlan, vlan_data in hosts[host]['vlans'].items():
                    portmap_sheet.cell(column=device_name_index + 1, row=row, value=device_name)
                    portmap_sheet.cell(column=interface_vlan_index + 1, row=row, value=vlan)
                    portmap_sheet.cell(column=interface_name_index + 1, row=row, value=f"Visio_VLAN_{num}")
                    portmap_sheet.cell(column=interface_description_index + 1, row=row, value=vlan_data['name'][:10])
                    num += 1
                    row += 1
        
        add_blank_line = True
        
        portmap_sheet.freeze_panes = "C2"
        
        # print(portmap_visio_data)
        with open('visio_interfaces.txt', 'a+') as f:
            f.write(portmap_visio_data)
        
    print(Fore.BLUE + f"\nsaving {filename}!" + Fore.RESET)
    # workbook.save(f"{filename[:-5]}_{summary_data['city'][:6]}.xlsx")
    workbook.save(filename)


def get_pepsi_cdp_info(directory, filename, hosts):
    print('Running get_pepsi_cdp_info')
    print(inspect.stack()[1].function)
    workbook = load_workbook(filename=filename)
    cdp_sheet = workbook["CDP"]
        
    column_names = []
    for cell in cdp_sheet[1]:
        column_names.append(cell.value)
    
    column_indexes = []
    for column in column_names:
        column_indexes.append(column_names.index(column))
    
    group_index = column_names.index("Group")
    local_device_index = column_names.index("Local Device")
    local_ip_index = column_names.index("Host Address")
    remote_device_index = column_names.index("Remote Device")
    remote_ip_index = column_names.index("IP Address")
    local_inface_index = column_names.index("Local Interface")
    remote_inface_index = column_names.index("Remote Interface")
    platform_index = column_names.index("Platform")
    capabilitiy_index = column_names.index("Capability")
    version_index = column_names.index("Version")
    holdtime_index = column_names.index("Holdtime")
    max_row = 2
    files = []
    cdp_neighbors = {}

    for row in cdp_sheet.iter_rows(max_col=20, min_row=2):
        if row[1].row > max_row:
            max_row = row[1].row

        local_device = str(row[local_device_index].value).split('.')[0]
        local_address = row[local_ip_index].value
        remote_device = str(row[remote_device_index].value).split('.')[0]
        remote_address = str(row[remote_ip_index].value).split(',')[0]
        holdtime = row[holdtime_index].value
        local_iface = normalize_interface_name(row[local_inface_index].value)
        remote_iface = normalize_interface_name(row[remote_inface_index].value)
        platform = row[platform_index].value
        capability = row[capabilitiy_index].value
        version = row[version_index].value
        group = row[group_index].value
        cdp_interface_data = {
            "local_address": local_address,
            "remote_device": remote_device,
            "remote_address": remote_address,
            "local_iface": local_iface,
            "remote_iface": remote_iface,
            "platform": platform,
            "capability": capability,
            "version": version
        }
        
        if local_device not in cdp_neighbors.keys():
            cdp_neighbors[local_device] = {}

        if 'neighbors' not in cdp_neighbors[local_device]:
            cdp_neighbors[local_device]['neighbors'] = [cdp_interface_data]
        else:
            cdp_neighbors[local_device]['neighbors'].append(cdp_interface_data)
                
    for host in hosts:
        if host not in cdp_neighbors.keys():
            continue

        for neighbor in cdp_neighbors[host]['neighbors']:
            for interface in hosts[host]['interfaces']:
                if neighbor['local_iface'].lower() == interface['name'].lower():
                    interface['neighbor'] = neighbor
  
    # remote_interface_name = remote_iface
    # remote_device_fqdn = remote_device
    # remote_device_name = extract_hostname_from_fqdn(remote_device_fqdn)
    # remote_interface = Interface(remote_interface_name, remote_device_name)
    # interface.neighbors.append(remote_interface)
    # host_interfaces[local_iface] = interface
    
    return hosts


def combine_backup_reports():
    from ttp import ttp
    import datetime
    directory = directory_check('BEReports', func='Backup Reports')
    date_template = """Report Date: {{ date }} {{ time }} {{ am_pm }}"""
    wb = Workbook()
    ws = wb.active
    duplicates = {}
    for root, dirs, files in os.walk(directory):
        for f in files:
            sitename = f.split()[-2].split('_')[0].lstrip('-') 
            print(f"Working on Site {sitename}!")
            
            # convert html files to pandas dataframe
            df_list = pd.read_html(io=f"{root}/{f}")
            
            # get the date of the report
            date_line = df_list[-2][0][0]
            parser = ttp(data=date_line, template=date_template)
            parser.parse()
            date_string = parser.result()[0][0]['date']
            date = datetime.datetime.strptime(date_string, "%m/%d/%Y")
            
            # check if report is newer than previous, report if duplicates
            if sitename not in duplicates.keys():
                duplicates[sitename] = []
            duplicates[sitename].append(date)
            
            if max(duplicates[sitename]) > date:
                print(f"\nSkipping {sitename} report with date {date} because a newer report with date {max(duplicates[sitename])} was previously found!")
                continue

            for df in df_list:
                if df.shape[0] >= 50:
                    df.columns = df.iloc[1]
                    df.drop(df.index[1], inplace=True)
                    df.dropna(how='all', inplace=True)

                    row = 2
                    sheet = wb.create_sheet(sitename)
                    sheet.title = sitename
                    # sheet.cell(row=1, column=1, value='Site')
                    for column in df.columns:
                        column_index = df.columns.get_loc(column)
                        sheet.cell(row=1, column=column_index + 2, value=column)

                    for index, session in df.iterrows():
                        for column in df.columns:
                            column_index = df.columns.get_loc(column)
                            sheet.cell(row=row, column=column_index + 1, value=str(session[column]))

                        row += 1
    wb.remove(ws)
    output_filename = 'BEReports_Combined.xlsx'
    print(f"Saving to {output_filename}!")
    wb.save(output_filename)
    wb.close()


def generate_asbuilt_data(summary_data):
    # update find and replace page
    filename = summary_data['filename']
    if 'Find_Replace' not in workbook.sheetnames:
        print('Did not find worksheet name Find_Replace, did you copy from the latest template?')
    else:
        find_replace_sheet = workbook['Find_Replace']
        _City = city.capitalize()
        _State = state.upper()
        _SiteID = site
        _ServerID = f"PF{bt_code.replace('-', '').upper()}"
        _Domain = 'ffa.pep.pvt'
        _SwitchID = site
        _BTDeviceID = f"PEP-{bt_code.upper()}"

        find_replace_sheet['B2'] = _City  # _City
        find_replace_sheet['B3'] = _State  # _State
        find_replace_sheet['B4'] = _SiteID  # _SiteID
        find_replace_sheet['B5'] = _ServerID  # _ServerID
        find_replace_sheet['B6'] = _SwitchID  # _SwitchID
        find_replace_sheet['B7'] = _BTDeviceID  # _BTDeviceID
        find_replace_sheet['B8'] = _Domain  # _Domain

        for subnet in summary_data['all_subnets']:
            
            if str(subnet['vlan']) == '600':
                _OTMANSubnet = '.'.join(subnet['network'].split('.')[:3])
                find_replace_sheet['B9'] = _OTMANSubnet
                find_replace_sheet['D9'] = subnet['vlan'] 
                find_replace_sheet['E9'] = subnet['description'] 
            
            elif str(subnet['vlan']) == '654':
                _DMZ1Subnet = '.'.join(subnet['network'].split('.')[:3])
                find_replace_sheet['B10'] = _DMZ1Subnet
                find_replace_sheet['D10'] = subnet['vlan'] 
                find_replace_sheet['E10'] = subnet['description'] 

            elif str(subnet['vlan']) == '601':
                _OTVMSubnet = '.'.join(subnet['network'].split('.')[:3])
                find_replace_sheet['B11'] = _OTVMSubnet
                find_replace_sheet['D11'] = subnet['vlan'] 
                find_replace_sheet['E11'] = subnet['description'] 
            
            elif 'OT Server' in subnet['description']:
                _OTServSubnet = '.'.join(subnet['network'].split('.')[:3])
                find_replace_sheet['B12'] = _OTServSubnet
                find_replace_sheet['D12'] = subnet['vlan'] 
                find_replace_sheet['E12'] = subnet['description'] 
            
            elif str(subnet['vlan']) == '65':
                _FWSubnet = '.'.join(subnet['network'].split('.')[:3])
                find_replace_sheet['B13'] = _FWSubnet
                find_replace_sheet['D13'] = subnet['vlan'] 
                find_replace_sheet['E13'] = subnet['description'] 

    # Update Table - OT MDF_IDC Rack Table
    description_index = 1
    hostname_index = 3
    ip_index = 4
    if 'Table - OT MDF_IDC Rack' not in workbook.sheetnames:
        print("Did not find worksheet name 'Table - OT MDF_IDC Rack', did you copy from the latest template?")
    else:
        ot_mdf_idc_sheet = workbook['Table - OT MDF_IDC Rack']
        for row in ot_mdf_idc_sheet.iter_rows(min_row=2, max_row=17, max_col=4):
            if 'Firewall 2' in row[description_index].value:
                ot_mdf_idc_sheet.cell(row=row[0].row, column=hostname_index, value=f"PEP-{bt_code.upper()}-FW01")
                ot_mdf_idc_sheet.cell(row=row[0].row, column=ip_index, value=f"{_FWSubnet}.210")
            elif 'Firewall 1' in row[description_index].value:
                ot_mdf_idc_sheet.cell(row=row[0].row, column=hostname_index, value=f"PEP-{bt_code.upper()}-FW01")
                ot_mdf_idc_sheet.cell(row=row[0].row, column=ip_index, value=f"{_FWSubnet}.209")
            
            elif 'iDMZ Switch' in row[description_index].value:
                ot_mdf_idc_sheet.cell(row=row[0].row, column=hostname_index, value=f"PEP-{bt_code.upper()}-DMZ01")
                ot_mdf_idc_sheet.cell(row=row[0].row, column=ip_index, value=f"{_DMZ1Subnet}.12")            
            
            elif 'OT Core Switch' in row[description_index].value:
                ot_mdf_idc_sheet.cell(row=row[0].row, column=hostname_index, value=f"{_SwitchID}38r01-OT")
                ot_mdf_idc_sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTMANSubnet}.1")

            elif 'IDC Server Switch' in row[description_index].value:
                ot_mdf_idc_sheet.cell(row=row[0].row, column=hostname_index, value=f"{_SwitchID}38s11-OT")
                ot_mdf_idc_sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTMANSubnet}.11")

            elif 'Host 4 iDRAC' in row[description_index].value:
                ot_mdf_idc_sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}IMH04")
                ot_mdf_idc_sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTMANSubnet}.14")             
                   
            elif 'Host 3 iDRAC' in row[description_index].value:
                ot_mdf_idc_sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}IHM03")
                ot_mdf_idc_sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTVMSubnet}.13")   

            elif 'Host 2 iDRAC' in row[description_index].value:
                ot_mdf_idc_sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}IHM02")
                ot_mdf_idc_sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTVMSubnet}.12")   

            elif 'Host 1 iDRAC' in row[description_index].value:
                ot_mdf_idc_sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}IHM01")
                ot_mdf_idc_sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTVMSubnet}.11")   

            elif 'PDU2' in row[description_index].value:
                ot_mdf_idc_sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}PDU02")
                ot_mdf_idc_sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTVMSubnet}.5")
            
            elif 'PDU1' in row[description_index].value:
                ot_mdf_idc_sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}PDU01")
                ot_mdf_idc_sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTVMSubnet}.4")  
            
            elif 'UPS' in row[description_index].value:
                ot_mdf_idc_sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}UPS01")
                ot_mdf_idc_sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTVMSubnet}.6")

    # Update Table - Cluster Details
    description_index = 1
    hostname_index = 1
    ip_index = 3
    sheet_name = 'Table - Cluster Details'
    if sheet_name not in workbook.sheetnames:
        print(f"Did not find worksheet name {sheet_name}, did you copy from the latest template?")
    else:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2, max_row=6, max_col=4):
            if "VMware vCenter Server" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}Vim01")
                sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTVMSubnet}.10")           
            
            elif "Host 1 ESXi (Management Host)" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}VMH01")
                sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTVMSubnet}.21") 
            
            elif "Host 2 ESXi (Cluster Host)" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}VMH02")
                sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTVMSubnet}.22")   

            elif "Host 3 ESXi (Cluster Host)" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}VMH03")
                sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTVMSubnet}.23")  
            
            elif "Host 4 ESXi (Cluster Host)" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}VMH04")
                sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTVMSubnet}.24")  

    # Update Table - Infrastructure VMs
    description_index = 1
    hostname_index = 1
    ip_index = 3
    sheet_name = 'Table - Infrastructure VMs'
    if sheet_name not in workbook.sheetnames:
        print(f"Did not find worksheet name {sheet_name}, did you copy from the latest template?")
    else:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2, max_row=7, max_col=4):
            if "Active Directory Domain Controller 1" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}WAD01")
                sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTServSubnet}.11")   
            
            elif "Active Directory Domain Controller 2" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}WAD02")
                sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTServSubnet}.12")   

            elif "Windows File Server" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}FIL01")
                sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTVMSubnet}.35")  

            elif "Backup Exec Server" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}WBAK01")
                sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTVMSubnet}.33")  

            elif "Virtual Support Engineer" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}WVSE01")
                sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTServSubnet}.76")  

            elif "Licensing Server" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}WLIC01")
                sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTVMSubnet}.34")  

    # Update Table - Application VMs
    match_object = "Site Prep Doc"
    print(match_object)
    filename = None
    for root, dirs, files in os.walk('./'):
        for f in files:
            if match_object in f and f.endswith('xlsx'):
                r1 = input(f"Use Site Prep Doc '{f}'? [y/n/q][default y]: ").lower().strip()
                if r1 == 'q':
                    exit()
                elif r1 == 'n':
                    continue
                else:
                    filename = f
                    break
    
    description_index = 2
    hostname_index = 1
    ip_index = 3
    app_vm_note_index = 4
    application_vms = []
    sheet_name = 'Table - Application VMs'
    if sheet_name not in workbook.sheetnames:
        print(f"Did not find worksheet name {sheet_name}, did you copy from the latest template?")
    else:
        if filename is not None:
            in_use_index = 0
            sp_ip_index = 1
            sp_naming_index = 2
            sp_description_index = 3
            notes_index = 12
            site_prep_wb = load_workbook(filename=filename, read_only=True)
            machine_creation_sheet = site_prep_wb['Machine Creation']
            for row in machine_creation_sheet.iter_rows(min_row=4, max_col=14, values_only=True):
                if row[in_use_index] is None:
                    continue
                
                if 'X' not in row[in_use_index].upper():
                    continue

                if 'Infrastructure Server' not in row[notes_index] and 'Application' in row[notes_index]:
                    server_data = {}
                    server_data['hostname'] = f"{_ServerID}{row[sp_naming_index]}"
                    server_data['ip'] = f"{_OTServSubnet}.{row[sp_ip_index].split('.')[-1]}"
                    server_data['description'] = row[sp_description_index]
                    server_data['notes'] = row[notes_index]
                    application_vms.append(server_data)
            
            site_prep_wb.close()
            
            row = 2
            sheet = workbook[sheet_name]
            for server in application_vms:
                sheet.cell(row=row, column=hostname_index, value=server['hostname'])
                sheet.cell(row=row, column=description_index, value=server['description'])
                sheet.cell(row=row, column=ip_index, value=server['ip'])
                sheet.cell(row=row, column=app_vm_note_index, value=server['notes'])
                row += 1
        
        else:
            print("Couldn't find a site prep doc to pull application server info, building standard list")
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=2, max_row=6, max_col=4):
                if "FactoryTalk Asset Center Server" in row[description_index].value:
                    sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}FTAC01")
                    sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTServSubnet}.24") 

                elif "FactoryTalk Asset Center Agent" in row[description_index].value:
                    sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}FTAG01")
                    sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTServSubnet}.25")
                
                elif "FactoryTalk Directory" in row[description_index].value:
                    sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}WFTD01")
                    sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTServSubnet}.23")  

                elif "Engineering Workstation" in row[description_index].value:
                    sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}WEWS01")
                    sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTServSubnet}.78")  

                elif "Microsoft SQL Server" in row[description_index].value:
                    sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}WSQL01")
                    sheet.cell(row=row[0].row, column=ip_index, value=f"{_OTServSubnet}.13") 

    # Update Table - SN & LK
    description_index = 1
    hostname_index = 1
    ip_index = 3
    sheet_name = 'Table - SN & LK'
    if sheet_name not in workbook.sheetnames:
        print(f"Did not find worksheet name {sheet_name}, did you copy from the latest template?")
    else:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2, max_row=17, max_col=4):
            if "Management Server" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}HM01")
            elif "Host 1" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}HM02")
            elif "Host 2" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}HM03")
            elif "Host 3" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}HM04")
            elif "IDC Server Switch" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"{_SwitchID}38s11-OT")
            elif "Core Switch" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"{_SwitchID}38r01-OT")
            elif "DMZ Switch 1" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"PEP-{bt_code.upper()}-DMZ01")
            elif "DMZ Switch 2" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"PEP-{bt_code.upper()}-DMZ02")
            elif "Firewall" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"PEP-{bt_code.upper()}-FW01")
            elif "IDF 12 Switch 1" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"{_SwitchID}36s12-OT")        
            elif "IDF 13 Switch 2" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"{_SwitchID}36s13-OT")     
            elif "UPS" in row[description_index].value:
                sheet.cell(row=row[0].row, column=hostname_index, value=f"{_ServerID}UPS01")


if __name__ == "__main__":
    
    # Start up Nornir!
    nr = InitNornir(config_file="config.yaml") 
    
    options = [
        "Generate Recomended IP Addresses From FFA Block on Summary Sheet",
        "Generate IP Address Information From Networks on Summary Sheet. VLAN and Description Required + (Subnet Mask + 1 or more of [starting, ending, gateway ip]) or (Network in CIDR Format)",
        "Generate Visio and Portmap data using portmaps/config files/cdp neighbors (Future option Show Techs or Live)",
        "Combine Symantec Backup Reports into and Excel File",
        "Generate As-Built Tables"
    ]

    response = ''

    first_run = True
    while response != 'q':
       
        # Get response from Main Menu then ask for parameters file if first run is True
        num = 1
        for option in options:
            print(f"{Fore.BLUE}({Fore.RESET}{num}{Fore.BLUE}){Fore.RESET}{Fore.GREEN + Style.BRIGHT} {option} {Fore.RESET}")
            num += 1
        
        print(f"{Fore.BLUE}({Fore.RESET}Q{Fore.BLUE}){Fore.RESET}{Fore.GREEN + Style.BRIGHT} Quit {Fore.RESET}")

        response = input("\n Please choose a option, [1-9][Q] to quit: ").lower()
        
        if first_run is True and response != 'q' and response != '4':
            first_run = False    

            output_filename = None
            filename = file_check(output_filename)

            workbook = load_workbook(filename=filename)
            summary_sheet = workbook["Summary"]
            bt_connection_sheet = workbook["BT Connection Table"]
            idmz_vlans_sheet = workbook["iDMZ VLANs"]
            transit_vlans_sheet = workbook["Transit VLANs"]

            bt_filename = "BT Site Codes.xlsx"
            # bt_filename = file_check(filename)

            bt_workbook = load_workbook(filename=bt_filename, read_only=True)
            bt_codes_sheet = bt_workbook["BT SITE ID's"]
    

            # city = input("Enter full city name (e.g., Jacksonville or StLouis):\n").capitalize()
            # state = input("Enter state abbreviation (e.g., FL or MO):\n ").upper()
            # bt_location = input("Enter BT location descriptor (USAM, CANA, etc.):\n").upper()
            # bt_code = int(input("Enter BT location code (505, 978, etc.):\n"))
            # ffa_block = input("Enter the assigned site FFA block network address (e.g., 11.29.208.0/20):\n")
            
            column_names = []
            for cell in summary_sheet[10]:
                column_names.append(cell.value)
            # print(column_names)
            
            # These column names must be in row two of Summary Sheet
            vlan_index = column_names.index("VLAN") 
            description_index = column_names.index("Description")
            gateway_index = column_names.index("Gateway Address")
            network_index = column_names.index("Network") 
            starting_index = column_names.index("Starting IP Address")
            ending_index = column_names.index("Ending IP Address")
            mask_index = column_names.index("Subnet Mask")
    
            state = summary_sheet['B3'].value
            city = summary_sheet['B4'].value
            bt_location = summary_sheet['B5'].value
            bt_code = summary_sheet['B6'].value
            address = summary_sheet['F3'].value
            pepsi_division = summary_sheet['F4'].value
            ffa_block_xls = summary_sheet['B8'].value
                        
            if not city:
                print("City Name was not found on summary sheet cell B4, please update, exiting..")
                exit()
            
            if bt_location is None or bt_code is None or address is None or pepsi_division is None:
                print("One of the items below is None, will try and update..\n")
                print(f"BT Location:  {bt_location}")
                print(f"BT Code: {bt_code}")
                print(f"Address: {address}")
                print(f"Pepsi Division: {pepsi_division}\n")
                # print(f"ffa_block {ffa_block_xls}")          
                
                bt_code_index = 0
                address_index = 9
                country_index = 1
                sector_index = 2
                division_index = 4
                
                site = None
                bt_sites_found = []
                
                for row in bt_codes_sheet.iter_rows(values_only=True, min_row=6, max_row=1953, max_col=10):
                    if city.lower() in row[address_index].lower():
                        # print(row[address_index])
                        bt_sites_found.append(row)
                
                if len(bt_sites_found) == 0:
                    print(f"No bt sites found when searching for city {city}. Please update the Summary page manually")
                
                elif len(bt_sites_found) == 1:
                    print(f"One Site found for the city {city} with address {bt_sites_found[0][address_index]}")
                    site = bt_sites_found[0]
                
                else:
                    print(Fore.RED + f"Multiple sites found for the city {city}")
                    num = 1
                    
                    for site in bt_sites_found:
                        print(f"{Fore.BLUE}({Fore.RESET}{num}{Fore.BLUE}){Fore.RESET} BT Site Code {site[0]}{Fore.BLUE} {site[address_index]}{Fore.RESET}")
                        num += 1
                    
                    site_choice = input("\nPlease choose the correct site address: ")
                    site = bt_sites_found[int(site_choice) - 1]
                    print(f"Site in {city} with BT Code {site[bt_code_index]} chosen!")

                if site:
                    bt_location = site[sector_index]
                    bt_code = site[bt_code_index]
                    address = site[address_index]
                    pepsi_division = site[division_index]
            
            if bt_location is None or bt_code is None or address is None or pepsi_division is None:
                print("Couldn't find summary information, please update bt location, bt code, address, division manually, exiting")
                exit()
            
            summary_data = {}
            summary_data['state'] = state
            summary_data['city'] = city
            summary_data['bt_location'] = bt_location
            summary_data['bt_code'] = bt_code
            summary_data['ffa_block_xls'] = ffa_block_xls
            summary_data['filename'] = filename
            
            all_subnets = []
            summary_data['all_subnets'] = all_subnets
            
            # Modify names
            site = f"pb{city.replace(' ', '')[:8]}{state}".lower() 
            fw01 = f"pep-{bt_code}-fw01".lower()
            fw02 = f"pep-{bt_code}-fw02".lower()
            
            summary_data['site'] = site
            summary_data['fw01'] = fw01
            summary_data['fw02'] = fw02
            
            # update excel summary page
            summary_sheet['B5'] = bt_location
            summary_sheet['B6'] = bt_code
            summary_sheet['A1'] = f"{site.capitalize()} Network Information"
            summary_sheet['B7'] = site           
            summary_sheet['F3'] = address
            summary_sheet['F4'] = pepsi_division
            
            # save workbook to be used
            workbook.save(filename)
            output_filename = f"{summary_data['bt_location']}Q_{city[:8].capitalize()}_{state.upper()}_IP_Address_Assignments_{strftime('Y%Y-M%m-D%d_H%H-M%M')}.xlsx"
            for row in summary_sheet.iter_rows(values_only=True, min_row=11, max_row=50):
                if row[vlan_index] is None or row[network_index] is None or row[network_index] == "xxx.xxx.xxx.xxx/xx" or row[network_index] == "N/A":
                    continue
                
                subnet = {}
                subnet['description'] = row[description_index]
                subnet['gateway'] = row[gateway_index]
                subnet['vlan'] = row[vlan_index]
                subnet['network'] = row[network_index]
                
                network = ipaddress.ip_network(subnet['network'])
                subnet['network_w_prefix'] = network.with_prefixlen  # Network

                subnet['netmask'] = str(network.netmask)  # Subnet Mask
                subnet['gateway'] = str(network[1])  # Gateway
                subnet['starting_ip'] = str(network[2])  # First IP After Gateway
                subnet['ending_ip'] = str(network[-2])  # Last IP

                all_subnets.append(subnet)

            # Get known hosts in target site
            nr_filtered = nr.filter(F(groups__contains=site))
            hosts = nr_filtered.inventory.hosts
            
            # Add New Hosts to inventory
            new_host_found = check_hosts_in_inventory(nr, hosts, summary_data)
            
            if new_host_found:
                input('Please Restart to Reload Inventory with newly found hosts, enter to exit')
                exit()
                # print(f"{Fore.BLUEprint(Fore.BLUE}Running On Site [{site}] {len(nr_filtered.inventory.hosts.keys())} "
                #       f"Hosts:  \n{(list(nr_filtered.inventory.hosts.keys()))}{Fore.RESET}")
                nr = InitNornir(config_file="config.yaml")
                nr_filtered = nr.filter(F(groups__contains=site))
            
            print(f"\n{Fore.BLUE + Style.BRIGHT}Running On {len(nr_filtered.inventory.hosts.keys())} Switches in Site {site} ")
            for host in hosts:
                print(Fore.BLUE + Style.BRIGHT + host + Fore.RESET)
            print('\n')
            
            # Check if hosts(switches) are in inventory
            # if len(hosts.keys()) < 3:
            #     print(len(hosts.keys()))
            # check_hosts_in_inventory(hosts, summary_data)
        
        if response == 'q':
            exit()
        
        elif response == '1':
            generate_networks_from_ffa(summary_data)
            workbook.save(output_filename)
            print(f"\n Saving to  {output_filename}!!!\n")
        
        elif response == '2':
            generate_ip_address_info(summary_data)
            workbook.save(output_filename)
            print(f"\n Saving to  {output_filename}!!!\n")

        elif response == '3':
            update_interfaces_from_config_files(hosts=hosts)
            # get_pepsi_cdp_info(directory='./Show_Techs', filename='Master_CDP.xlsx', hosts=hosts)
            generate_portmap_data(hosts, summary_data)
            workbook.save(output_filename)
            print(f"\n Saving to  {output_filename}!!!\n")
        
        elif response == '4':
            combine_backup_reports()
        
        elif response == '5':
            # print(summary_data)
            generate_asbuilt_data(summary_data)
            workbook.save(output_filename)
            print(f"\n Saving to  {output_filename}!!!\n")
       
        else:
            print(f"\n{Fore.RED + Style.BRIGHT}The Input [{response}] is Invalid!{Fore.RESET}") 

    workbook.close()
