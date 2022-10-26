import collections
import difflib
import inspect
import ipaddress
import itertools as it
import os
import re
import shutil
import threading
import time
from time import strftime

import matplotlib.pyplot as plt
import networkx as nx
import nmap
from colorama import Fore, Style
from mac_vendor_lookup import MacLookup
from nornir.core.filter import F
from nornir.plugins.functions.text import print_result
from nornir.plugins.tasks import text
from nornir.plugins.tasks.networking import (netmiko_save_config,
                                             netmiko_send_command,
                                             netmiko_send_config, tcp_ping)
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from ruamel.yaml import YAML
from tabulate import tabulate

from constants import INTERFACE_NAME_RE, NORMALIZED_INTERFACES
from interface import Interface

LOCK = threading.Lock()

def check_hosts_in_inventory(nr, hosts, summary_data):

    directory = "./Show Tech/Configs"
    site = summary_data['site']
    new_hosts = {}
    for root, dirs, files in os.walk(directory):
        for f in files:
            new_host_data = {}
            hostname = None
            ip = None            
            
            with open(f"{root}/{f}", 'r', errors='replace') as a:
                config_lines = a.readlines()
            
            for index, line in enumerate(config_lines):
                if line.startswith('hostname'):
                    # print(line)
                    hostname = line.split()[1]
                    # print(f"Hostname [{hostname}]")
                    
                    if hostname not in f:
                        os.rename(f"{root}/{f}", f"{root}/{hostname}.cfg")
                        print(Fore.RED + f"\nSwitch {hostname} found in config file {f} but is not in the filename,\
                         \nThe script resolves this by renaming config files to hostname found in config name but please investigate why." + Fore.RESET)
                    
                # This is not perfect below, create logic for perfect end of file searching, something like, while index + range <= len(config_lines)
                elif 'vlan260' in line.lower():
                    # print(len(config_lines) - index)
                    if (len(config_lines) - index) > 3:
                        for i in range(1, 3):
                            if 'address' in config_lines[index + i].lower():
                                # print(config_lines[index + i])
                                ip = config_lines[index + i].split()[2]
               
                elif 'vlan261' in line.lower():
                    if (len(config_lines) - index) > 3:
                        for i in range(1, 3):
                            if 'address' in config_lines[index + i].lower():
                                # print(config_lines[index + i])
                                ip = config_lines[index + i].split()[2]   
                
                elif 'vlan600' in line.lower():
                    if (len(config_lines) - index) > 3:
                        for i in range(1, 3):
                            if 'address' in config_lines[index + i].lower():
                                # print(config_lines[index + i])
                                ip = config_lines[index + i].split()[2]                
            
            if hostname is None:
                print(f"No Hostname found in file {f}")
            else:
                if site[:7] not in hostname and not hostname.startswith('pep-'):
                    print(f"{Fore.RED + Style.BRIGHT}\nSite Name {site} does not match device name {hostname} found in config files. This is probably due to mismatch between show tech's site name found on summary page.")
                    print(f"Please Resolve, then restart!\n {Fore.RESET}")
                    exit()
                if hostname not in hosts.keys():
                    if ip is not None:
                        new_host_data = {}
                        new_hosts[hostname] = new_host_data
                        new_host_data['hostname'] = ip
                        new_host_data["groups"] = [summary_data['site'], 'cisco_ios']

                    else:
                        print(Fore.RED + f"Couldn't find IP for Host {hostname} in config file, add manually")
                    # path for configuration masters
            
            master_config = os.path.join(f"change_control/master_config_files/{summary_data['site']}/{hostname}.cfg")
            
            if not os.path.isdir(os.path.dirname(master_config)):
                os.makedirs(os.path.dirname(master_config))
                # copying current config to master config if not found.
            
            if not os.path.isfile(master_config):
                print(f"master config not found for {hostname} creating one from show run cmd_output")
                with open(master_config, "w") as f:
                    for line in config_lines:
                        f.write(line)
    print("\n")
    new_host_found = False
    import ipdb
    for host in new_hosts:
        if host not in hosts.keys():
            new_host_found = True
            print(Fore.BLUE + f"Switch {host} wasn't in inventory, adding to site {site}" + Fore.RESET)
            hosts[host] = new_hosts[host]
    
    if new_host_found:
        write_hosts_data_to_inventory_file(hosts)
    
    return new_host_found

  
def update_interfaces_from_config_files(hosts):
    
    configs_dir = "./Show Tech/Configs"
    # configs_dir = directory_check(directory=configs_dir, func='CONFIGS')
    hosts_not_found = hosts.copy()
    no_config = []
    
    if not os.path.isdir(os.path.dirname(configs_dir)):
        print(f"{configs_dir}  is no good")
    
    for root, dirs, files in os.walk(configs_dir):
        for config_file in files:
            for host in hosts:
                host_data = hosts[host]
                
                if host.lower() not in config_file.lower():
                    continue
                
                # print(hosts_not_found.keys())
                del hosts_not_found[host]
                with open(f"{root}/{config_file}", "r") as f:
                    config_lines = f.readlines()
    
                host_interfaces = collections.OrderedDict()
                vlans = {}
                description = ""
                host_data["interfaces"] = host_interfaces
                host_data['vlans'] = vlans
                
                for index, line in enumerate(config_lines):
                    try:
                        if line.startswith('vlan'):
                            vlan = line.lstrip('vlan').strip()
                            vlan_name = config_lines[index + 1].lstrip(' name ').strip()
                            vlans[vlan] = {'name': vlan_name}
                        if line.startswith("interface"):
                            interface_data = {}
                            interface = normalize_interface_name(line.strip().lstrip("interface ").lower())
                            
                            host_interfaces[interface] = interface_data
                            
                            interface_config_lines = []
                            interface_data["config"] = interface_config_lines
                            
                            for i in range(1, 6):
                                if config_lines[index + i].startswith(("!", "end", "exit")):
                                    # print(f'breaking index {index + i}')
                                    break
                                elif config_lines[index + i].strip().startswith("description"):
                                    description = config_lines[index + i].strip().lstrip("description")
                                    interface_data["description"] = description
                                    continue
                                
                                elif config_lines[index + i].startswith(" switchport access vlan"):
                                    vlan = config_lines[index + i].lstrip(" switchport access vlan ")
                                    interface_data['vlan'] = vlan
                                    # print(vlan)
                                
                                elif not config_lines[index + i].strip() == "":
                                    interface_config_lines.append(config_lines[index + i].strip())
                            
                            if len(interface_config_lines) == 0:
                                interface_data.pop("config", None)
                            
                            if len(host_interfaces[interface]) == 0:
                                del host_interfaces[interface]
                    
                    except IndexError:
                        pass
    
    if len(hosts_not_found) > 0:
        for host in hosts_not_found:
            print(Fore.RED + f"A Config file for {host} wasn't found" + Fore.RESET)


def file_check(filename, match_object):
    """Searches for previously created parameters file if filename is None and Match Object[STR] is provided

    Args:
        filename (String): Name of a file to check if exists
        match_object (String): Unique part of a filename to search for

    Returns:
        [String]: [Valid filename found]
    """    
    valid_file = False
    while valid_file is not True:
        # print(match_object)
        if filename is None:
            for root, dirs, files in os.walk('./'):
                for f in files:
                    if match_object in f and f.endswith('xlsx') and valid_file is not True:
                        r1 = input(f"{Fore.BLUE + Style.BRIGHT}\nUse Parameters File '{f}'? [y/n/q][default y]: {Fore.RESET}").lower().strip()
                        
                        if r1 == 'q':
                            exit()
                        
                        elif r1 == 'n':
                            continue
                        
                        elif r1 == '' or r1 == 'y':
                            filename = f
                            valid_file = True
                            break

                        else:
                            print(f"The Input [{r1}] isn't valid")

            
            if valid_file is not True:
                filename = file_check(filename=None, match_object="_Template")
                valid_file = True
                                 
        else:
            r1 = input(f"Use Parameters File '{filename}'? [y/n/q][default y]: ").lower()
            if r1 == 'q':
                exit()
            
            elif r1 == 'n':
                r2 = input("Enter the Parameters Filename: ")
                filename = r2
            
            if not os.path.isfile(filename):
                print(Fore.RED + f"{filename} not found" + Fore.RESET)
            
            else:
                print(Fore.BLUE + Style.BRIGHT + f"{filename} is valid, continuing...\n" + Fore.RESET)
                valid_file = True
    
    return filename


def directory_check(directory, func):
    valid_directory = False
    while valid_directory is not True:
        r1 = input(f"\nLook in '{directory}' for {func} files? [y/n/q][default y]: ").lower()
        if r1 == 'n':
            r2 = input("Enter the Directory: ")
            directory = r2
            if os.path.isdir(directory):
                valid_directory = True
        elif r1 == 'q':
            exit()
        
        if not os.path.isdir(directory):
            print(Fore.RED + f"{directory} not found" + Fore.RESET)
        else:
            print(Fore.BLUE + f"{directory} is valid, continuing...\n" + Fore.RESET)
            valid_directory = True
    return directory


def normalize_interface_name(interface_name: str):
    """ Normalizes interface name

    For example, GigabitEthernet1 is converted to Gi0/1,
    TenGigabitEthernet1/1  is converted to Te1/1
    """
    match = INTERFACE_NAME_RE.search(interface_name)
    if match:
        int_type = match.group("interface_type")
        # print(int_type)
        normalized_int_type = int_type[0:2:]
        int_num = match.group("interface_num")
        # print(int_num)
        return normalized_int_type + int_num
    raise ValueError(
        f"Does not recognize {interface_name} as an interface name")


def write_hosts_data_to_inventory_file(hosts):
    path = "./inventory/hosts.yaml"
    groups_path = "./inventory/groups.yaml"
    
    yaml = YAML()
    yaml.preserve_quotes = True
    with open(path) as f:
        parsed_yaml = yaml.load(f)
    
    with open(groups_path) as f:
        g_parsed_yaml = yaml.load(f)

    # for host in hosts:
    #     print(type(host))
    #     if "groups" in hosts[host]:
    #         print(hosts[host]['groups'])
    site = None
    for host in hosts:
        if host not in parsed_yaml:
            parsed_yaml[host] = hosts[host]
            site = hosts[host]['groups'][0]

    if site not in g_parsed_yaml:
        # print(g_parsed_yaml)
        site_data = {}
        site_data['data'] = {}
        site_data['data']['site'] = site
        g_parsed_yaml[site] = site_data

    with open(path, "w") as f:
        yaml.dump(parsed_yaml, f)
    
    with open(groups_path, "w") as f:
        yaml.dump(g_parsed_yaml, f)


def normalize_interface_type(interface_type: str) -> str:
    """Normalizes interface type

    For example, G is converted to GigabitEthernet, Te is converted to TenGigabitEthernet
    """
    int_type = interface_type.strip().lower()
    for norm_int_type in NORMALIZED_INTERFACES:
        if norm_int_type.lower().startswith(int_type):
            return norm_int_type
    return int_type


def parse_switch_configs(summary_data):

    directory = "./Show Tech/Configs"
    new_hosts = {}
    site = summary_data['site']
    networks = summary_data['networks']
    
    # look for config files in Show Tech's
    for root, dirs, files in os.walk(directory):
        for f in files:
            new_host_data = {}
            hostname = None
            ip = None
            vlan = None
            subnet = None           
            
            with open(f"{root}/{f}", 'r', errors='replace') as a:
                config_lines = a.readlines()
            
            for index, line in enumerate(config_lines):
                if line.startswith('hostname'):
                    # print(line)
                    hostname = line.split()[1]
                    # print(f"Hostname [{hostname}]")
                    
                    if hostname not in f:
                        os.rename(f"{root}/{f}", f"{root}/{hostname}.cfg")
                        print(Fore.RED + f"\nSwitch {hostname} found in config file {f} but is not in the filename,\
                         \nThe script resolves this by renaming config files to hostname found in config name but please investigate why." + Fore.RESET)
                
                # Look for SVI's and Switch Management IP's
                # This is not perfect below, create logic for perfect end of file searching, something like, while index + range <= len(config_lines)
                
                # Pepsi Management IP's
                pepsi_management_vlans = ('vlan260', 'vlan261', 'vlan600', 'vlan601')
                if any(mgmt_vlan in line.lower() for mgmt_vlan in pepsi_management_vlans):
                    if (len(config_lines) - index) > 3:
                        network = {}
                        vlan = line.lower().lstrip("interface vlan")
                        network['vlan'] = vlan
                        
                        for i in range(1, 3):
                            if config_lines[index + i].lower().startswith(" ip address"):
                                if 'no ip address' not in config_lines[index + i].lower():
                                    
                                    ip = config_lines[index + i].split()[2] 
                                    subnet = config_lines[index + i].split()[3]

                                    network_obj = ipaddress.ip_network(f"{ip}/{subnet}", strict=False)
                                    network['network_obj'] = network_obj
                                    networks.append(network)
                            
                            elif config_lines[index + i].lower().startswith(" description"):
                                description = config_lines[index + i].lstrip(" description")
                                network['name'] = description
                        
                        if 'network_obj' in network:
                            networks.append(network)
                
                # All Other SVI's
                elif line.lower().startswith("interface vlan"):
                    if (len(config_lines) - index) > 3:
                        network = {}
                        vlan = line.lower().lstrip("interface vlan")
                        network['vlan'] = vlan
                        
                        for i in range(1, 3):
                            if config_lines[index + i].lower().startswith(" ip address"):
                                if 'no ip address' not in config_lines[index + i].lower():
                                    
                                    ip = config_lines[index + i].split()[2] 
                                    subnet = config_lines[index + i].split()[3]

                                    network_obj = ipaddress.ip_network(f"{ip}/{subnet}", strict=False)
                                    network['network_obj'] = network_obj
                                    networks.append(network)
                            
                            elif config_lines[index + i].lower().startswith(" description"):
                                description = config_lines[index + i].lstrip(" description")
                                network['name'] = description
                        
                        if 'network_obj' in network:
                            networks.append(network)

            if hostname is None:
                print(f"No Hostname found in file {f}")
            
            else:
                if ip is not None:
                    valid_ip = False
                    while not valid_ip or ip == 'q':                       
                        try:

                            ip_obj = ipaddress.ip_address(ip)
                            valid_ip = True
                            print(f" {Fore.BLUE + Style.BRIGHT}{ip} is a valid IP Address{Fore.RESET}")
                        
                        except ValueError:
                            print(f"ip [{ip}] for {hostname} is Invalid!")
                            ip = input(f"IP {ip} found for host {hostname} is invalid please enter it now. exx 10.10.10.10: ").lower()
                            if ip == 'q':
                                exit()
                                        
                    print(f"{Fore.BLUE + Style.BRIGHT}\nUsing IP {ip} for {hostname} no need to change this for now if wrong SVI")

                    new_host_data = {}
                    new_hosts[hostname] = new_host_data
                    new_host_data['hostname'] = ip
                    new_host_data["groups"] = [summary_data['site'], 'cisco_ios']

                else:
                    valid_ip = False
                    while not valid_ip or ip == 'q':
                        ip = input(Fore.RED + f"Couldn't find IP for Host {hostname} please enter it now. ex 10.10.10.10: ").lower()
                        
                        try:
                            ip_obj = ipaddress.ip_address(ip)
                            valid_ip = True
                            print(f"{Fore.BLUE + Style.BRIGHT} IP Address {ip} is a valid IP Address{Fore.RESET}")
                        
                        except ValueError:
                            print(f"{ip} is Invalid!")
                            
            # path for configuration masters
            master_config = os.path.join(f"change_control/master_config_files/{summary_data['site']}/{hostname}.cfg")
            
            if not os.path.isdir(os.path.dirname(master_config)):
                os.makedirs(os.path.dirname(master_config))
                # copying current config to master config if not found.
            
            if not os.path.isfile(master_config):
                print(f"master config not found for {hostname} creating one from show run cmd_output")
                with open(master_config, "w") as f:
                    for line in config_lines:
                        f.write(line)

    return new_hosts


def get_pepsi_cdp_info(directory, filename, hosts):
    print('Running get_pepsi_cdp_info')
    print(inspect.stack()[1].function)
    workbook = load_workbook(filename=filename)
    cdp_sheet = workbook["CDP"]
        
    column_names = []
    for cell in cdp_sheet[1]:
        column_names.append(cell.value)
    
    column_indexes = []
    for column in column_names:
        column_indexes.append(column_names.index(column))
    
    group_index = column_names.index("Group")
    local_device_index = column_names.index("Local Device")
    local_ip_index = column_names.index("Host Address")
    remote_device_index = column_names.index("Remote Device")
    remote_ip_index = column_names.index("IP Address")
    local_inface_index = column_names.index("Local Interface")
    remote_inface_index = column_names.index("Remote Interface")
    platform_index = column_names.index("Platform")
    capabilitiy_index = column_names.index("Capability")
    version_index = column_names.index("Version")
    holdtime_index = column_names.index("Holdtime")
    max_row = 2
    files = []
    cdp_neighbors = {}

    for row in cdp_sheet.iter_rows(max_col=20, min_row=2):
        if row[1].row > max_row:
            max_row = row[1].row

        local_device = str(row[local_device_index].value).split('.')[0]
        local_address = row[local_ip_index].value
        remote_device = str(row[remote_device_index].value).split('.')[0]
        remote_address = str(row[remote_ip_index].value).split(',')[0]
        holdtime = row[holdtime_index].value
        local_iface = normalize_interface_name(row[local_inface_index].value)
        remote_iface = normalize_interface_name(row[remote_inface_index].value)
        platform = row[platform_index].value
        capability = row[capabilitiy_index].value
        version = row[version_index].value
        group = row[group_index].value
        cdp_interface_data = {
            "local_address": local_address,
            "remote_device": remote_device,
            "remote_address": remote_address,
            "local_iface": local_iface,
            "remote_iface": remote_iface,
            "platform": platform,
            "capability": capability,
            "version": version
        }
        
        if local_device not in cdp_neighbors.keys():
            cdp_neighbors[local_device] = {}

        if 'neighbors' not in cdp_neighbors[local_device]:
            cdp_neighbors[local_device]['neighbors'] = [cdp_interface_data]
        else:
            cdp_neighbors[local_device]['neighbors'].append(cdp_interface_data)
                
    for host in hosts:
        if host not in cdp_neighbors.keys():
            continue

        for neighbor in cdp_neighbors[host]['neighbors']:
            for interface in hosts[host]['interfaces']:
                if neighbor['local_iface'].lower() == interface['name'].lower():
                    interface['neighbor'] = neighbor
  
    # remote_interface_name = remote_iface
    # remote_device_fqdn = remote_device
    # remote_device_name = extract_hostname_from_fqdn(remote_device_fqdn)
    # remote_interface = Interface(remote_interface_name, remote_device_name)
    # interface.neighbors.append(remote_interface)
    # host_interfaces[local_iface] = interface
    
    return hosts


def generate_portmap_data(hosts, summary_data):
    """ Gets info from config files and arp files to create portmap data

    Args:
        hosts ([list]): list of nornir host objects
        summary_data ([type]): all of the data on the summary page 
    """    

    print(Fore.BLUE + Style.BRIGHT + "\nStarting Portmap" + Fore.RESET)
    
    # load the excel document 
    filename = summary_data['filename']
    workbook = load_workbook(filename=filename)
    
    portmap_template_sheet = workbook["Port Map Template"]
    column_names = []
    for cell in portmap_template_sheet[1]:
        column_names.append(cell.value)
    # print(column_names)

    # These column names must be in row two of Port Map Template Sheet
    device_name_index = column_names.index("Name")
    device_ip_index = column_names.index("IP Address")
    
    interface_name_index = column_names.index("Interface")
    interface_vlan_index = column_names.index("VLAN")
    interface_status_index = column_names.index("Status")
    interface_config_index = column_names.index("Config Lines")
    interface_neighbor_index = column_names.index("CDP Neighbor")
    interface_description_index = column_names.index("Interface Description")

    visio_portmap_index = column_names.index("Visio Portmap")
    connected_device_description_index = column_names.index("Connected Device Description")
    connected_device_remote_interface_index = column_names.index("Remote Interface")
    multi_mac_index = column_names.index("Multi-MAC")
    connected_ip_index = column_names.index("Connected Device IP")
    connected_mac_index = column_names.index("Connected Device MAC")
    connected_mac_vendor_index = column_names.index("Connected Device Mac Vendor")
    vlan_template_sheet = workbook["VLAN Template"]
    tcolumn_names = []
    
    for cell in vlan_template_sheet[4]:
        tcolumn_names.append(cell.value)

    # These column names must be in row 4 of vlan Template Sheet data sheet 
    t_mac_index = tcolumn_names.index("MAC Address") + 1
    t_mac_vendor_index = tcolumn_names.index("MAC Vendor") + 1
    t_switch_index = tcolumn_names.index("Switch") + 1
    t_interface_index = tcolumn_names.index("Switchport") + 1
    t_interface_description_index = tcolumn_names.index("Interface Description") + 1

    
    directory = './Show Tech/MAC_IP'
    arps = {}
    
    # Get [IP - Mac - VLAN] Relationship for all known IP's From MAC_IP's

    print(f'{Fore.BLUE + Style.BRIGHT}Getting Known IPs from MAC_IP Files{Fore.RESET}')
    if not os.path.isdir(os.path.dirname(directory)):
        print(Fore.RED + f"{directory}  was not found!" + Fore.RESET)
    
    for root, dirs, files in os.walk(directory):
        for f in files:
            # print(f"root {root}, dirs {dirs}, file {f}")
            # print(f"found arp file {f}")
            with open(f"{root}/{f}", 'r') as a:
                arp_lines = a.readlines()
            for line in arp_lines:
                if line.startswith('Internet'):
                    # print(line)
                    ip = line.split()[1]
                    vlan = line.split()[-1].lstrip('Vlan')
                    arp_entry = re.search(r"(\w+)\s+(\d+\.\d+\.\d+\.\d+)\s+(\d+)\s+(\w+.\w+.\w+)\s+(\w+)\s+(\S+)", line,)
                    if arp_entry:

                        device_mac = arp_entry.group(4)
                        # print(device_mac)
                        if ip in arps.keys() and arps[ip]['mac'] == device_mac:
                            continue
                        arps[ip] = {}     
                        arps[ip]['mac'] = device_mac
                        arps[ip]['vlan'] = vlan 
    

    # Get [HOST - INT - IP - MAC - VLAN] Relationship from MAC_IP's
    hosts_missing_arp_files = hosts.copy()
    hostname = None
    for root, dirs, files in os.walk(directory):
        for f in files:

            with open(f"{root}/{f}", 'r') as a:
                arp_lines = a.readlines()
               
            for ip in arps.keys():
                for line in arp_lines:
                    
                    # set the active host if hostname found in line
                    for host in hosts:
                        if host in line:
                            hostname = host
                            # delete host from list for debugging
                            if host in hosts_missing_arp_files.keys():
                                del hosts_missing_arp_files[host]
                    if hostname is None:
                        continue

                    if arps[ip]['mac'] in line:

                        iface = line.split()[-1]
                        
                        # if iface.lower().startswith('p'):
                        #     continue
                        # elif iface.lower().startswith('v'):
                        #     continue
                        
                        if iface.lower() in hosts[hostname]['interfaces'].keys():
                            # print(f"{iface} found on {host}")
                            if 'connected_devices' not in hosts[hostname]['interfaces'][iface.lower()]:
                                hosts[hostname]['interfaces'][iface.lower()]['connected_devices'] = {}
                            hosts[hostname]['interfaces'][iface.lower()]['connected_devices'][ip] = {'mac': arps[ip]['mac'], 'vlan': arps[ip]['vlan'], 'mac_vendor': MacLookup().lookup(arps[ip]['mac'])}
    
    for host in hosts_missing_arp_files:
        print(Fore.RED + f"Did not find a MAC_IP File for {host}" + Fore.RESET)    
    
    # Portmap - Per VLAN ->
    print(Fore.BLUE + "\nStarting Portmap - Per VLAN" + Fore.RESET)
    missing_ip_sheet = {}
    
    # loop through Hosts
    for host in hosts:
        num_connected = 0
        
        if 'interfaces' not in hosts[host]:
            print(f"no interfaces on {host}")
            continue
        
        for interface in hosts[host]['interfaces']:
            if 'connected_devices' not in hosts[host]['interfaces'][interface]:
                # print(f"No connected devices on {host} {interface}")
                continue
            else:
                # print(f"{host} {interface} Has Connected {len(hosts[host]['interfaces'][interface]['connected_devices'])} Devices")
                pass
            
            connected_devices = hosts[host]['interfaces'][interface]['connected_devices']
            
            # print(len(connected_devices))
            if len(connected_devices) >= 1:
                for ip in connected_devices:
                    ip_vlan = connected_devices[ip]['vlan']
                    # print(ip_vlan)
                                        
                    num_connected += 1
                    sheet_found = False
                    ip_found = False
                    vlan_sheetname = None
                    
                    for sheetname in workbook.sheetnames:
                        
                        if str(connected_devices[ip]['vlan']) in sheetname:
                            # print('vlan')
                            sheet_found = True
                            vlan_sheetname = sheetname
                            break
                    
                    if sheet_found is False:
                        print(f"No VLAN Sheet found for VLAN {ip_vlan}")
                        # sheet = workbook.copy_worksheet(workbook['VLAN Template'])
                        # sheet.title = (f"VLAN {ip_vlan}")
                    
                    else:
                        sheet = workbook[sheetname]
                    
                    for row in sheet.iter_rows(min_row=5):
                        if ip == row[0].value:

                            # print(f"{Fore.BLUE}{ip} found on {host} {interface} adding to {sheetname} sheet")
                            ip_found = True
                            sheet.cell(row=row[0].row, column=t_mac_index, value=connected_devices[ip]['mac'].upper())
                            sheet.cell(row=row[0].row, column=t_mac_vendor_index, value=connected_devices[ip]['mac_vendor'])
                            sheet.cell(row=row[0].row, column=t_interface_index, value=interface.capitalize())
                            sheet.cell(row=row[0].row, column=t_switch_index, value=host.upper())

                            if 'description' in hosts[host]['interfaces'][interface]:
                                sheet.cell(row=row[0].row, column=t_interface_description_index, value=hosts[host]['interfaces'][interface]['description'])
                                
                    if sheet_found is False:
                        print(f"No sheet found for {ip} on vlan {connected_devices[ip]['vlan']}")
                        pass
                    
                    if ip_found is False:
                        missing_ip_sheet[connected_devices[ip]['vlan']] = ip
             
        print(Fore.RED + f"{host} Has {num_connected} Connected Devices" + Fore.RESET)
    # Portmap - Per VLAN - END ^

    # Portmap - Per Switch ->
    print(Fore.BLUE + "\nStarting Portmap - Per Switch" + Fore.RESET)
    visio_row = 150
    
    for host in hosts:
        num_connected = 0
        
        sheet_found = False
        add_blank_line = False
        
        host_data = hosts[host]
        device_name = host
        device_ip = hosts[host]['hostname']
        portmap_lines = []
        portmap_visio_data = ""
        
        row = 2
        for sheetname in workbook.sheetnames:
            if host in sheetname:
                sheet_found = True
                portmap_sheet = workbook[sheetname] 
        
        if not sheet_found:
            portmap_sheet = workbook.copy_worksheet(portmap_template_sheet)
            portmap_sheet.title = f"{host} Port Map"

        # Loop Interfaces
        if 'interfaces' not in hosts[host]:
            print(f"no interfaces on {host} this is most likey and error!")
            continue
        
        for iface_name, interface in host_data["interfaces"].items():
                       
            portmap_visio_data = ""
            
            interface_vlan = None
            interface_config = None
            interface_description = None
            
            interface_neighbor = None
            interface_neighbor_remote_interface = None
            
            connected_device_name = None
                            
            connected_device_description = None
            connected_device_remote_interface = None
            
            multi_mac = False

            connected_devices = None
            
            if 'config' in interface.keys():
                interface_config = "\n".join(interface['config'])
            
            if 'description' in interface.keys():
                interface_description = interface['description']

            if 'vlan' in interface.keys():
                interface_vlan = interface['vlan']

            if 'neighbor' in interface.keys():
                interface_neighbor = interface['neighbor']['remote_device']
                connected_device_ip = interface['neighbor']['remote_address']
                connected_device_remote_interface = interface['neighbor']['remote_iface']
            
            # loop for connected devices
            if 'connected_devices' in interface.keys():
                
                connected_devices = interface['connected_devices']

                if len(connected_devices) >= 1:             
                    # Create Data for Pepsi Portmap descriptions
                    portmap_visio_data += f"\n{iface_name.capitalize()}"
                    for ip in connected_devices:
                        connected_device_name = ip
                        connected_device_description = ip
                        portmap_visio_data += "\n"
                        portmap_visio_data += f"{ip}\n"
                        
                        if 'description' in interface.keys():
                            portmap_visio_data += f"{interface['description'].strip()}\n"
                        
                        portmap_visio_data += f"{connected_devices[ip]['mac'].upper()}\n"
                                    
                        if not connected_device_name:
                            connected_device_name = interface_neighbor

            if add_blank_line:
                # row += 1
                portmap_sheet.insert_rows(row)
                row += 1
                add_blank_line = False
            
            if connected_devices is None:
                portmap_sheet.insert_rows(row)
                portmap_sheet.cell(column=device_name_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                portmap_sheet.cell(column=device_name_index + 1, row=row, value=device_name)
                portmap_sheet.cell(column=device_ip_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                portmap_sheet.cell(column=device_ip_index + 1, row=row, value=device_ip)
                
                if iface_name:
                    portmap_sheet.cell(column=interface_name_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                    portmap_sheet.cell(column=interface_name_index + 1, row=row, value=iface_name.capitalize())
                
                if interface_vlan:
                    portmap_sheet.cell(column=interface_vlan_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                    portmap_sheet.cell(column=interface_vlan_index + 1, row=row, value=int(interface_vlan))
                
                if interface_config:
                    portmap_sheet.cell(column=interface_config_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                    portmap_sheet.cell(column=interface_config_index + 1, row=row, value=interface_config.strip())
                
                portmap_sheet.cell(column=interface_neighbor_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                portmap_sheet.cell(column=interface_neighbor_index + 1, row=row, value=interface_neighbor)
                                
                portmap_sheet.cell(column=interface_description_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                portmap_sheet.cell(column=interface_description_index + 1, row=row, value=interface_description)
                
                if portmap_visio_data:
                    portmap_sheet.cell(column=visio_portmap_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                    portmap_sheet.cell(column=visio_portmap_index + 1, row=row, value=portmap_visio_data.strip())
                    
                portmap_sheet.cell(column=connected_device_description_index + 1, row=row, value=connected_device_description)
                
                if connected_device_remote_interface:
                    portmap_sheet.cell(column=connected_device_remote_interface_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                    portmap_sheet.cell(column=connected_device_remote_interface_index + 1, row=row, value=connected_device_remote_interface.capitalize())
                
                portmap_sheet.row_dimensions[row].height = 15
                row += 1
            
            elif len(connected_devices) == 1:
                for ip in connected_devices:

                    portmap_sheet.insert_rows(row)
                    portmap_sheet.cell(column=connected_ip_index + 1, row=row, value=ip)
                    portmap_sheet.cell(column=connected_mac_index + 1, row=row, value=connected_devices[ip]['mac'].upper())
                    portmap_sheet.cell(column=connected_mac_vendor_index + 1, row=row, value=connected_devices[ip]['mac_vendor'])
                    portmap_sheet.cell(column=device_name_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                    portmap_sheet.cell(column=device_name_index + 1, row=row, value=device_name)
                    portmap_sheet.cell(column=device_ip_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                    portmap_sheet.cell(column=device_ip_index + 1, row=row, value=device_ip)
                    
                    if iface_name:
                        portmap_sheet.cell(column=interface_name_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                        portmap_sheet.cell(column=interface_name_index + 1, row=row, value=iface_name.capitalize())
                    
                    if interface_vlan:
                        portmap_sheet.cell(column=interface_vlan_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                        portmap_sheet.cell(column=interface_vlan_index + 1, row=row, value=int(interface_vlan))
                    
                    if interface_config:
                        portmap_sheet.cell(column=interface_config_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                        portmap_sheet.cell(column=interface_config_index + 1, row=row, value=interface_config.strip())
                    
                    portmap_sheet.cell(column=interface_neighbor_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                    portmap_sheet.cell(column=interface_neighbor_index + 1, row=row, value=interface_neighbor)
                                    
                    portmap_sheet.cell(column=interface_description_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                    portmap_sheet.cell(column=interface_description_index + 1, row=row, value=interface_description)
                    
                    if portmap_visio_data:
                        portmap_sheet.cell(column=visio_portmap_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                        portmap_sheet.cell(column=visio_portmap_index + 1, row=row, value=portmap_visio_data.strip())
                        
                    portmap_sheet.cell(column=connected_device_description_index + 1, row=row, value=connected_device_description)
                    
                    if connected_device_remote_interface:
                        portmap_sheet.cell(column=connected_device_remote_interface_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                        portmap_sheet.cell(column=connected_device_remote_interface_index + 1, row=row, value=connected_device_remote_interface.capitalize())
                    
                    portmap_sheet.row_dimensions[row].height = 15
                    row += 1

            elif len(connected_devices) >= 2:
                for ip in connected_devices:

                    portmap_sheet.insert_rows(row)
                    portmap_sheet.cell(column=multi_mac_index + 1, row=row, value='X')
                    portmap_sheet.cell(column=connected_ip_index + 1, row=row, value=ip)
                    portmap_sheet.cell(column=connected_mac_index + 1, row=row, value=connected_devices[ip]['mac'].upper())
                    portmap_sheet.cell(column=connected_mac_vendor_index + 1, row=row, value=connected_devices[ip]['mac_vendor'])
                    portmap_sheet.cell(column=device_name_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                    portmap_sheet.cell(column=device_name_index + 1, row=row, value=device_name)
                    portmap_sheet.cell(column=device_ip_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                    portmap_sheet.cell(column=device_ip_index + 1, row=row, value=device_ip)
                    
                    if iface_name:
                        portmap_sheet.cell(column=interface_name_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                        portmap_sheet.cell(column=interface_name_index + 1, row=row, value=iface_name.capitalize())
                    
                    if interface_vlan:
                        portmap_sheet.cell(column=interface_vlan_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                        portmap_sheet.cell(column=interface_vlan_index + 1, row=row, value=int(interface_vlan))
                    
                    if interface_config:
                        portmap_sheet.cell(column=interface_config_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                        portmap_sheet.cell(column=interface_config_index + 1, row=row, value=interface_config.strip())
                    
                    portmap_sheet.cell(column=interface_neighbor_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                    portmap_sheet.cell(column=interface_neighbor_index + 1, row=row, value=interface_neighbor)
                                    
                    portmap_sheet.cell(column=interface_description_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                    portmap_sheet.cell(column=interface_description_index + 1, row=row, value=interface_description)
                    
                    if portmap_visio_data:
                        portmap_sheet.cell(column=visio_portmap_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                        portmap_sheet.cell(column=visio_portmap_index + 1, row=row, value=portmap_visio_data.strip())
                        
                    portmap_sheet.cell(column=connected_device_description_index + 1, row=row, value=connected_device_description)
                    
                    if connected_device_remote_interface:
                        portmap_sheet.cell(column=connected_device_remote_interface_index + 1, row=row).alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='left', vertical='top')
                        portmap_sheet.cell(column=connected_device_remote_interface_index + 1, row=row, value=connected_device_remote_interface.capitalize())
                    
                    portmap_sheet.row_dimensions[row].height = 15
                    row += 1
            
        num = 1
        if 'vlans' in hosts[host].keys():
            for vlan, vlan_data in hosts[host]['vlans'].items():
                portmap_sheet.cell(column=device_name_index + 1, row=row, value=device_name)
                portmap_sheet.cell(column=interface_vlan_index + 1, row=row, value=int(vlan))
                portmap_sheet.cell(column=interface_name_index + 1, row=row, value=f"Visio_VLAN_{num}")
                portmap_sheet.cell(column=interface_description_index + 1, row=row, value=vlan_data['name'][:10])
                num += 1
                row += 1

        
        add_blank_line = True
        
        portmap_sheet.freeze_panes = "D2"
        
        
    print(Fore.BLUE + Style.BRIGHT + f"\nSaving {filename}!\n" + Fore.RESET)
    # workbook.save(f"{filename[:-5]}_{summary_data['city'][:6]}.xlsx")
    # workbook.remove(vlan_template_sheet)
    # workbook.remove(portmap_template_sheet)
    workbook.save(filename)


def get_routing_info(task):
    r = task.run(netmiko_send_command, command_string="show ip route ospf", use_textfsm="True")
    # b = task.run(netmiko_send_command, command_string="show ip bgp summary", use_textfsm="True")
    # c = task.run(netmiko_send_command, command_string="show ip bgp", use_textfsm="True")
    
    # print(r.result)

    task.host["show-ip-route"] = r.result

    route_data = task.host["show-ip-route"]

    network_list = []
    mask_list = []
    next_hop_list = []
    next_hop_int_list = []
    protocol_list = []

    headers = ["OSPF Learned Prefixes", "Next Hop IP", "Next Hop Interface"]
    for number in route_data:
        network_list.append(f"{number['network']}\{number['mask']}")
        mask_list.append(number['mask'])
        next_hop_list.append(number['nexthop_ip'])
        next_hop_int_list.append(number['nexthop_if'])
        protocol_list.append(number['protocol'])
    

def display_infograph(headers, data):
    # Incomplete as of 6/11/20
     
    list1 = []
    list2 = []
    list3 = []
    list4 = []
    list5 = []
    
    # LOCK.acquire()
    # for number in data:
    #     for header in headers:
    #         for i in range(4):
    #             list1.append(f"{number['network']}\{number['mask']}")
    #     list2.append(number['mask'])
    #     list3.append(number['nexthop_ip'])
    #     list4.append(number['nexthop_if'])
    #     list5.append(number['protocol'])
    # print("\n")
    # print(Fore.GREEN + Style.BRIGHT + "*" * 111)
    # print(Fore.YELLOW + f"OSPF Routing Information for {task.host}")
    # print(Fore.GREEN + Style.BRIGHT + "*" * 111)
    # table = it.zip_longest(network_list, next_hop_list,next_hop_int_list)
    # print(tabulate(table, headers=headers, tablefmt="psql"))
    # print(Fore.MAGENTA + Style.BRIGHT + "*" * 111)
    # LOCK.release()


def get_cdp_info(task):
    c = task.run(netmiko_send_command, command_string="show cdp neighbor detail", use_textfsm="True")
    
    task.host["cdp-neigh"] = c.result
    
    cdp_data = task.host["cdp-neigh"]

    neigh_list = []
    neigh_ip_list = []
    neigh_plat_list = []
    local_port_list = []
    remote_port_list = []
    
    headers = ["Neighbor", "Neighbor IP", "Neighbor Platform", "Local Port", "Remote Port"]
    for number in cdp_data:
        neigh_list.append(number['destination_host'])
        neigh_ip_list.append(number['management_ip'])
        neigh_plat_list.append(number['platform'])
        local_port_list.append(number['local_port'])
        remote_port_list.append(number['remote_port'])
    
    LOCK.acquire()
    print("\n")
    print(Fore.GREEN + Style.BRIGHT + "*" * 111)
    print(Fore.YELLOW + f"CDP Information for {task.host}")
    print(Fore.GREEN + Style.BRIGHT + "*" * 111)
    table = it.zip_longest(neigh_list, neigh_ip_list, neigh_plat_list, local_port_list, remote_port_list)
    print(tabulate(table, headers=headers, tablefmt="psql"))
    print(Fore.MAGENTA + Style.BRIGHT + "*" * 111)
    LOCK.release()

def extract_hostname_from_fqdn(fqdn: str):
    """Extracts hostname from fqdn-like string

    For example, R1.cisco.com -> R1,  sw1 -> sw1"
    """
    return fqdn.split(".")[0]


def list_files(directory):
    files = next(os.walk(directory))[2]
    files.sort()
    return files


def delete_files_in_dir(directory):
    for the_file in os.listdir(directory):
        file_path = os.path.join(directory, the_file)
        try:
            if os.path.isfile(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print(e)


def ssh_capable(task):
    result = task.run(
        task=netmiko_send_command,
        name="SSH Capable",
        enable=True,
        use_textfsm=True,
        command_string="show version",
    )
    print(result.result[0])
    ssh_vers = ("k9", "k2", "k8")
    if any(ver in result.result[0]["running_image"] for ver in ssh_vers):
        print(result.result[0]["running_image"])
        print(f"{task.host.name} is SSH Capable")
        with open("ssh_capable.txt", "a") as f:
            f.write(f"{task.host.name} \\n")


def change_control(nr_filtered):
    # get the running config from device

    hosts = nr_filtered.inventory.hosts

    nr_filtered.run(task=change_control_show_run)

    config_diff(hosts)


def change_control_show_run(task):
    task.run(task=ping)
    if task.host["ping"] is True:
        # print(f"Running on {task.host.name}")
        r = task.run(
            task=netmiko_send_command,
            name="Show Run Netmiko",
            command_string="show run",
            enable=True,
        )
        task.host["running_config"] = r.result
    else:
        print(f"host [{task.host.name}] failed to respond to ping, Skipping")


def config_diff(hosts):
    for host in hosts:
        host_data = hosts[host]
        host_ip = hosts[host].hostname

        if "running_config" in host_data:

            # print(f"\nstarting config diff for device {host}")
            host_data["changes_detected"] = None
            running_config = host_data["running_config"]

            # path for configuration masters
            master_config = os.path.join(
                f"change_control/master_config_files/{host_data['site']}/{host}.txt"
            )
            if not os.path.isdir(os.path.dirname(master_config)):
                os.makedirs(os.path.dirname(master_config))

            backups = os.path.join(
                f"change_control/backups/{host_data['site']}/{host}/{host}_{strftime('%Y-%m-%d_%M')}.txt"
            )
            if not os.path.isdir(os.path.dirname(backups)):
                os.makedirs(os.path.dirname(backups))

            current_config = os.path.join(f"change_control/temp/{host}.txt")
            if not os.path.isdir(os.path.dirname(current_config)):
                os.makedirs(os.path.dirname(current_config))

            # path for change reports
            change_report = os.path.join(
                f"change_control/per_device_change_reports/{host_data['site']}/{strftime('%Y-%m-%d_%M')}/{host}.txt"
            )
            if not os.path.isdir(os.path.dirname(change_report)):
                os.makedirs(os.path.dirname(change_report))

            with open(current_config, "w") as f, open(backups, "w") as b:
                f.write(running_config)
                b.write(running_config)

            # copying current config to master config if not found.
            if not os.path.isfile(master_config):
                print(
                    f"master config not found for {host} creating one from show run cmd_output"
                )
                with open(master_config, "w") as f:
                    f.write(running_config)

            # Start ndiff
            with open(master_config, "r") as m, open(current_config, "r") as c:
                diff = list(difflib.ndiff(m.readlines(), c.readlines()))

            # Iterating over the diff and extracting the differences and saving to master differences_report_file
            # Using try/except to catch and ignore any IndexError exceptions that might occur.
            with open(change_report, "w") as report_file:
                try:
                    for index, line in enumerate(diff):
                        if (
                            line.startswith("- ! Last configuration change") or 
                            line.startswith("+ ! Last configuration change")
                            or line.startswith("+ ! NVRAM config")
                            or line.startswith("- Current configuration")
                            or line.startswith("+ Current configuration")
                            or line.startswith("- ntp clock-period")
                            or line.startswith("- !")
                            or line.startswith("+ !")
                        ):
                            
                            continue

                        elif (line.startswith("- ") and diff[index + 1].startswith(("?", "+")) is False):
                            report_file.write(f"\nWas in old version, not there anymore: \n\n{line} \n-------\n\n")

                        elif line.startswith("+ ") and diff[index + 1].startswith("?") is False:
                            report_file.write(f"\nWas not in old version, is there now: \n\n...\n{line}...\n\n-------\n")

                        elif (line.startswith("- ") and diff[index + 1].startswith("?") and diff[index + 2].startswith("+ ") and diff[index + 3].startswith("?")):
                            report_file.write(f"\nChange detected here: \n\n{line + diff[index + 1]}{diff[index + 2]}{diff[index + 3]}\n-------\n")

                        elif (line.startswith("- ") and diff[index + 1].startswith("+") and diff[index + 2].startswith("? ")):
                            report_file.write(f"\nChange detected here: \n\n{line + diff[index + 1]}{diff[index + 2]}\n-------\n")

                        else:
                            pass
                except IndexError:
                    pass

            if os.path.getsize(change_report) == 0:
                # print(f"Removing {host} diff file because its size was {os.path.getsize(change_report)}")
                os.remove(change_report)
            else:
                print(
                    f"Size of diff file for {os.path.getsize(change_report)} is {host} setting changes_detected to True")

                with open(change_report, "r") as f:
                    host_data["diff"] = f.readlines()

                host_data["changes_detected"] = True

                with open(master_config, 'w') as m:
                    m.write(running_config)


def get_subnets(subnets_file):
    subnets = []
    with open(subnets_file, "r") as file:
        subnets_list = file.readlines()
        for subnet in subnets_list:
            subnets.append(subnet.rstrip("\n") + "/24")
        # print(subnets)
    return subnets


def create_master_report(task):
    hostname = task.host.name
    master_report_file = os.path.join(
        f"change_control/all_device_changes/{strftime('%Y-%m-%d')}_master_report.txt"
    )
    if not os.path.isdir(os.path.dirname(master_report_file)):
        os.makedirs(os.path.dirname(master_report_file))
    # Reading the report file and writing to the master_report_file.
    if task.host["changes_detected"] is True:
        with open(master_report_file, "a+") as master_report:
            master_report.write(
                "\n*** Changes Detected for Device: " + hostname + " ***\n"
            )
            master_report.writelines(task.host["diff"])
    # copyfile(device_current_config, master_config_file)
    # print("Master Configuration updated with current config for " + hostname)
    elif task.host["changes_detected"] is False:
        with open(master_report_file, "r") as master_report, open(
            "temp_master.txt", "w"
        ) as temp_master:
            temp_master.write(f"\n*** No changes detected on: {hostname}")
            temp_master.write(master_report.read())
        os.rename("temp_master.txt", master_report_file)


def template_deploy(task):
    render_template_results = task.run(
        task=text.template_file,
        name="render configs",
        template="baseline.j2",
        path=f"./templates/{task.host.platform}",
    )

    task.host["config"] = render_template_results.result

    send_template_results = task.run(
        task=netmiko_send_config,
        name="send_template",
        config_commands=task.host["config"].split("\n")
    )

    print_result(send_template_results)

    task.run(task=netmiko_save_config, name="Save Config", confirm=True)


def update_ntp(task):
    print(f"ntp server {task.host['ntp']}")
    task.run(
        task=netmiko_send_config,
        name="Send config  Netmiko",
        config_commands=[f"ntp server {task.host['ntp']}"],
    )


def update_documentation(nr_filtered, filename, update_switch_description=False):
    
    # # save a master config using the change_control function
    change_control(nr_filtered=nr_filtered)
    
    # # pull data out of master config file as host_data["interfaces"]
    update_interfaces_from_config_files(nr_filtered.inventory.hosts)
    
    # # get interface status as host_data["show ip interfaces brief"]
    nr_filtered.run(task=send_command, command="show ip int brief", use_textfsm=True, write_output_to_file=False, print_output=False)
    
    # get cdp information as host_data[show cdp neighbors detail"]
    nr_filtered.run(task=get_cdp_neighbors)

    # load the excel document
    from openpyxl import load_workbook
    workbook = load_workbook(filename=filename)
    sheet = workbook["Physical Connections"]
    column_names = []
    for cell in sheet[2]:
        column_names.append(cell.value)
    # print(column_names)

    # These column names must be in row two of Physical Connections sheet
    device_name_index = column_names.index("Device Name")
    ip_address_index = column_names.index("IP Address")
    status_index = column_names.index("Status")
    config_lines_index = column_names.index("Config Lines")
    vlan_index = column_names.index("VLAN")
    interface_index = column_names.index("Port")
    description_index = column_names.index("Description")
    remote_port_index = column_names.index("Remote Port")
    connected_device_ip_index = column_names.index("Connected Device IP")
    remote_switch_index = column_names.index("Remote Switch")
    # print(workbook.sheetnames)
    max_row = 3
    
    for host in nr_filtered.inventory.hosts:
        
        host_data = nr_filtered.inventory.hosts[host]           
        host_ip = nr_filtered.inventory.hosts[host].hostname

        # print(host_data.keys())
        if "show ip int brief" in host_data:
            # print(f"show ip int brief found in host_data for [{host}]")

            add_blank_line = True
            for interface in host_data["show ip int brief"]:
                interface_name = normalize_interface_name(interface['intf'])
                interface_status = interface['status']
                confg_lines = None
                interface_vlan = None
                interface_description = None
                interface_found_in_excel = False
                interface_neighbor = None
                interface_neighbor_remote_interface = None
                interface_neighbor_ip = None
                
                # Use interface information found in config file to update config_lines and description
                if "interfaces_from_config" in host_data:
                    
                    if interface_name in host_data['interfaces_from_config']:
                        print(f"[{interface_name}] for host [{host}] found in config file")
                        
                        interface_info_from_config_file = host_data['interfaces_from_config'][interface_name]
                        if 'config' in interface_info_from_config_file:
                            config_lines = "\n".join(interface_info_from_config_file['config'])
                            
                        if 'description' in interface_info_from_config_file and interface_description is None:
                            interface_description = interface_info_from_config_file['description']
                            # print(interface_description)

                if "show cdp neighbors detail" in host_data:
                    # print(f"cdp neighbors found in host_data for {host}")
                    neighbors = host_data["show cdp neighbors detail"]
    
                    for neighbor in neighbors:
                        if interface_name == neighbors[neighbor]["local_interface"]:
                            # print(f"CDP Neighbor [{neighbor}] discovered on [{host}] int [{interface_name}]")
                            interface_neighbor = neighbor
                            interface_neighbor_remote_interface = neighbors[neighbor]["remote_interface"]
                            interface_neighbor_ip = neighbors[neighbor]["ip"]
                                        
                # max_row is used to find the last line with data
                for row in sheet.iter_rows(max_col=20, min_row=2):
                    if row[1].row > max_row:
                        max_row = row[1].row

                    # if interface is found in excel update it
                    interface_from_excel = normalize_interface_name(str(row[interface_index].value))
                    if row[device_name_index].value == host and interface_from_excel == interface_name:
                        # print(f"[{interface_name}] for host [{host}] found in excel, updating")
                        
                        interface_found_in_excel = True
                        # Update excel  
                        sheet.cell(column=status_index + 1, row=row[1].row, value=interface_status)
                        sheet.cell(column=vlan_index + 1, row=row[1].row, value=interface_vlan)
                        sheet.cell(column=config_lines_index + 1, row=row[1].row, value=config_lines)
                        sheet.cell(column=remote_switch_index + 1, row=row[1].row, value=interface_neighbor)
                        sheet.cell(column=remote_port_index + 1, row=row[1].row, value=interface_neighbor_remote_interface)
                        sheet.cell(column=connected_device_ip_index + 1, row=row[1].row, value=interface_neighbor_ip)
                        if row[description_index].value is None:
                            sheet.cell(column=description_index + 1, row=row[1].row, value=interface_description)

                if not interface_found_in_excel:
                    # print(f"[{interface_name}] for host [{host}] not found in excel")
                    
                    add_row_num = max_row + 1
                    if add_blank_line:

                        sheet.insert_rows(add_row_num)
                        max_row += 1
                        add_blank_line = False

                    add_row_num = max_row + 1
                    sheet.insert_rows(add_row_num)
                    sheet.cell(column=device_name_index + 1, row=add_row_num, value=host)
                    sheet.cell(column=description_index + 1, row=add_row_num, value=interface_description)
                    sheet.cell(column=status_index + 1, row=add_row_num, value=interface_status)
                    sheet.cell(column=vlan_index + 1, row=add_row_num, value=interface_vlan)
                    sheet.cell(column=interface_index + 1, row=add_row_num, value=interface_name)
                    sheet.cell(column=ip_address_index + 1, row=add_row_num, value=host_ip)
                
            if "show cdp neighbors detail" in host_data: 
                neighbors = host_data["show cdp neighbors detail"]
                for neighbor in neighbors:
                    nr_cdp_neighbor_search = nr_filtered.filter(F(hostname=neighbors[neighbor]["ip"]) | F(name=neighbor))
                    nr_cdp_neighbor_search_hosts = nr_cdp_neighbor_search.inventory.hosts
                    if not bool(nr_cdp_neighbor_search_hosts):
                        print(f"CDP Neighbor [{neighbor}] with ip [{neighbors[neighbor]['ip']}] is not in inventory")

                        add_row_num = max_row + 1
                        sheet.insert_rows(add_row_num)
                        add_row_num += 1
                        sheet.insert_rows(add_row_num)

                        sheet.cell(column=device_name_index + 1, row=add_row_num, value=neighbor)
                        sheet.cell(column=interface_index + 1, row=add_row_num, value=neighbors[neighbor]["remote_interface"])
                        sheet.cell(column=remote_switch_index + 1, row=add_row_num, value=host)
                        sheet.cell(column=remote_port_index + 1, row=add_row_num, value=neighbors[neighbor]["local_interface"])
                        sheet.cell(column=connected_device_ip_index + 1, row=add_row_num, value=host_ip)
                        sheet.cell(column=ip_address_index + 1, row=add_row_num, value=neighbors[neighbor]["ip"])



def add_hosts_from_configs(nr_filtered, configs_dir):
    config_names = list_files(configs_dir)
    hosts = nr_filtered.inventory.hosts
    new_hosts = {}
    for config in config_names:
        hostname = None
        mgmt_ip = None
        gateway = None
        config_path = f"{configs_dir}/{config}"
        if os.path.isfile(config_path):
            print(f"found config {config}")
            host_interfaces = {}
            with open(config_path, "r") as f:
                config_lines = f.readlines()
            for index, line in enumerate(config_lines):
                try:
                    if line.startswith("hostname"):
                        hostname = line.strip().lstrip("hostname ")
                        print(hostname)
                    elif line.startswith(" ip address"):
                        mgmt_ip = line.strip().lstrip(" ip address ").split()[0]
                        print(mgmt_ip)
                    elif line.startswith("ip default-gateway"):
                        gateway = line.strip().lstrip("ip default-gateway ")
                        print(gateway)
                except IndexError:
                    pass
        else:
            print(f"somethings wrong with the path [{config_path}] ")
            
        if hostname and mgmt_ip and gateway:
            new_host_data = {}
            new_hosts[hostname] = new_host_data
            new_host_data["hostname"] = mgmt_ip
            site = input("Please enter a site name: ")
            new_host_data["groups"] = [site, 'cisco_ios']
            master_config_path = f"./change_control/master_config_files/{site}/{hostname}.cfg"
            if not os.path.isdir(os.path.dirname(master_config_path)):
                os.makedirs(os.path.dirname(master_config_path))
            shutil.copyfile(config_path, master_config_path)

    for host in new_hosts:
        hosts[host] = new_hosts[host]
    
    write_hosts_data_to_inventory_file(hosts)


def discover_add_hosts(nr_filtered, site):
    nr_filtered.run(task=get_cdp_neighbors)
    hosts = nr_filtered.inventory.hosts
    new_hosts = {}
    
    for host in hosts:
        host_data = hosts[host]
        host_ip = hosts[host].hostname

        if "show cdp neighbors detail" in host_data:
            neighbors = host_data["show cdp neighbors detail"]

            for neighbor in neighbors:
                # print(neighbor)
                nr_cdp_neighbor_search = nr_filtered.filter(F(hostname=neighbors[neighbor]["ip"]) | F(name=neighbor))
                nr_cdp_neighbor_search_hosts = nr_cdp_neighbor_search.inventory.hosts

                if not bool(nr_cdp_neighbor_search_hosts):
                    print(f"new host found via cdp neighbors connected to [{host}] on port [{neighbors[neighbor]['local_interface']}] with name [{neighbor}] and ip [{neighbors[neighbor]['ip']}]")

                    new_host_data = {}
                    new_hosts[neighbor] = new_host_data
                    new_host_data["hostname"] = neighbors[neighbor]['ip']
                    new_host_data["groups"] = [site, 'cisco_ios_telnet']

    for host in new_hosts:
        hosts[host] = new_hosts[host]

    write_hosts_data_to_inventory_file(hosts)


def get_cdp_neighbors(task):
    """can only be run on devices that support textfsm ex ios, call write_hosts_data after"""
    result = task.run(
        task=netmiko_send_command,
        enable=True,
        name="Send Command Netmiko",
        command_string="show cdp neighbors detail",
        delay_factor=5,
        use_textfsm=True,
    )
    
    neighbors = {}
    task.host["show cdp neighbors detail"] = neighbors
    if len(result.result) >= 1:
        for interface_info in result.result:
            local_interface_name = normalize_interface_name(interface_info["local_port"])
            remote_interface_name = normalize_interface_name(interface_info["remote_port"])
            remote_device_fqdn = interface_info["destination_host"]
            remote_device_name = extract_hostname_from_fqdn(remote_device_fqdn)
            remote_ip = interface_info["management_ip"]
            neighbors[remote_device_name] = {
                "local_interface": local_interface_name,
                "remote_interface": remote_interface_name,
                "ip": remote_ip,
            }
    else:
        print(f"No Neighbors found on [{task.host.name}]")
    # print(neighbors)


def ping(task):
    result = task.run(task=tcp_ping, ports=[22, 23])
    if True not in result.result.values():
        # print(f"Ping failed for {task.host.name}'with ip {task.host.hostname} :-( ")
        task.host["ping"] = False
        with open("nornir_ping_fail.log", "a+") as f:
            f.write(
                f"{strftime('%Y-%m-%d_%H-%M')}  {task.host.name} failed ping test\n")
    else:
        # print(f"Ping passed for {task.host.name}'with ip {task.host.hostname}! ")
        task.host["ping"] = True
        with open("nornir_ping_pass.log", "a+") as f:
            f.write(
                f"{strftime('%Y-%m-%d_%H-%M')}  {task.host.name} passed ping test!\n")


def send_command(task, command, use_textfsm=False, write_output_to_file=True, print_output=True):
    # task.run(task=ping)
    # if task.host.data["ping"] is True:
    r = task.run(
        task=netmiko_send_command,
        name=f"{task.host.name}: {command}",
        command_string=command,
        enable=True,
        use_textfsm=use_textfsm
    )
    if "Invalid input detected " in r.result:
        print(f"Invalid input detected on host [{task.host.name}] input sent [{command}]")
        # print_result(r)
    else:
        task.host[command] = r.result
        if write_output_to_file:
            task.run(task=write_output, command=command)
        if print_output:
            print_result(r)


def send_config(task, config, save_config=True):
    task.run(task=ping)
    if task.host.data["ping"] is True:
        r = task.run(
            task=netmiko_send_config, name="send_config", config_commands=config
        )
        if "Invalid input detected " in r.result:
            print(f"Invalid input detected on host [{task.host.name}]")
            print(r.result)
            pass
        else:
            print_result(r)
            task.host["results"] = r.result
            if save_config:
                save_result = task.run(task=netmiko_save_config, name="Save Config", confirm=True)
                # print_result(save_result)


def write_output(task, command="None Provided"):
    hostname = task.host.name
    output = task.host[command]
    command_output = os.path.join(
        f"command_output/{strftime('Y%Y-M%m-D%d_%H-%M')}/{task.host['site']}/{command.replace(' ', '_').replace('|', '').replace('.', '_')[:10]}/{hostname}.txt"
    )
        
    if not os.path.isdir(os.path.dirname(command_output)):
        os.makedirs(os.path.dirname(command_output))

    with open(command_output, "w+") as command_file:
        command_file.write(str(output))


def write_interfaces_to_inventory(task):
    path = "/home/saroot/Network_Device_Management/inventory/hosts.yaml"
    yaml = YAML()
    yaml.preserve_quotes = False
    with open(path) as f:
        parsed_yaml = yaml.load(f)

    if "data" not in parsed_yaml[task.host.name]:
        parsed_yaml[task.host.name]["data"] = {}
    parsed_yaml[task.host.name]["data"]["interfaces"] = task.host.data["interfaces"]
    with open(path, "w") as f:
        yaml.dump(parsed_yaml, f)


def populate_arp_table(task, devices):
    """POPULATE ARP/CAM TABLE by Pining known devices"""

    for device in devices:
        # print(f'running device_to_int_relationship on {device.name}')
        if device.hostname == task.host.hostname:
            print(f"{device.name} is current switch, skipping")
            continue
        else:
            task.run(
                task=netmiko_send_command,
                name="ping",
                command_string=f"ping {device.hostname}",
                enable=True,
            )
    result = task.run(
        task=netmiko_send_command,
        name="show arp",
        command_string="show arp",
        enable=True,
        use_textfsm=True,
    )
    
    # print(result.result)
    for line in result.result.splitlines():
        arp_entry = re.search(
            r"(\w+)\s+(\d+\.\d+\.\d+\.\d+)\s+(\d+)\s+(\w+.\w+.\w+)\s+(\w+)\s+(\S+)",
            line,
        )
        if arp_entry:
            for device in devices:
                if arp_entry.group(2) == device.hostname:
                    device.mac = arp_entry.group(4)
                    # print(f"{device.name} was found in mac table as {device.mac}")

    result = task.run(
        task=netmiko_send_command,
        name="show mac addresses",
        command_string="show mac address-table",
        enable=True,
        use_textfsm=True,
    )

    task.host["show_mac"] = result


def device_int_relationship(hosts, devices):
    # Find if there is a device to port association on this switch and add to local_devices list
    for host in hosts:
        print(f"Running on {len(devices)} devices")
        local_devices = []
        for device in devices:
            if "show_mac" not in host:
                continue
            result = host["show_mac"]
            for mac_entry in result.result:
                if "neighbors" in host.data:
                    for key in host["neighbors"].keys():
                        if (
                            mac_entry["destination_port"]
                            in host["neighbors"][key]["local_interface"]
                        ):
                            print(
                                f"Excluding Mac entry for {mac_entry['destination_port']} because its a neighbor"
                            )
                            break
                if device.mac == mac_entry["destination_address"]:
                    device.interface = mac_entry["destination_port"]
                    device.vlan = mac_entry["vlan"]
                    print({k: v for k, v in device.__dict__.items() if v is not None})
                    device_data = {
                        "hostname": device.hostname,
                        "mac": device.mac,
                        "vlan": device.vlan,
                    }
                    if device.os:
                        device_data["os"] = device.os
                    print(
                        f"device to int association found {device.name} on {host.name} port {device.interface}"
                    )
                    local_devices.append(device)
                    if device.interface in host["interfaces"]:
                        print(
                            f"interfaces {device.interface} already in inventory for device {host.name}"
                        )
                        if "connected_devices" in host["interfaces"][device.interface]:
                            print(
                                f"'connected_devices' for {host.name} and {device.interface} already in inventory"
                            )
                            if (
                                device.name
                                in host["interfaces"][device.interface][
                                    "connected_devices"
                                ]
                            ):
                                continue
                            else:
                                print(device.name)
                                print(device.interface)
                                host["interfaces"][device.interface][
                                    "connected_devices"
                                ][device.name] = device_data
                        else:
                            print(
                                f"'connected_devices' for {host.name} and {device.interface} not in inventory"
                            )
                            host["interfaces"][device.interface][
                                "connected_devices"
                            ] = {}
                            host["interfaces"][device.interface]["connected_devices"][
                                device.name
                            ] = device_data

                    else:
                        print(
                            f"interface {device.interface} for {host.name} not in inventory"
                        )
                        host["interfaces"][device.interface] = {}
                        host["interfaces"][device.interface]["connected_devices"] = {}
                        host["interfaces"][device.interface]["connected_devices"][
                            device.name
                        ] = device_data

            host.data.pop("show_mac", None)
        # removing devices found locally
        devices = [x for x in devices if x not in local_devices]


def find_dict_key(key, var):
    """Finds key in dict var and returns value(s)"""
    if hasattr(var, "items"):
        for k, v in var.items():
            if k == key:
                yield v
            if isinstance(v, dict):
                for result in find_dict_key(key, v):
                    yield result
            elif isinstance(v, list):
                for d in v:
                    for result in find_dict_key(key, d):
                        yield result


def update_neighbors(task):
    task.run(task=ping)
    if task.host.data['ping'] is True:
        result = task.run(task=netmiko_send_command,
                          enable=True,
                          name="Send Command Netmiko",
                          command_string='show cdp neighbors detail',
                          delay_factor=5,
                          use_textfsm=True)

        device_name = task.host.name
        host_interfaces = {}
        # print(type(result.result))
        task.host.data["interfaces"] = host_interfaces
        for interface_info in result.result:
            interface_name = interface_info['local_port']
            
            interface = Interface(interface_name, device_name)
            
            remote_interface_name = interface_info["remote_port"]
            remote_device_fqdn = interface_info["destination_host"]
            
            remote_device_name = extract_hostname_from_fqdn(remote_device_fqdn)
            remote_interface = Interface(remote_interface_name, remote_device_name)
            
            interface.neighbors.append(remote_interface)
            host_interfaces[interface.name] = interface


def build_graph(hosts):
    edge_labels = [{}, {}]
    
    links = set([interface.link_from_neighbors() for host in hosts for interface in host.data["interfaces"].values()])
    print(links)
    
    graph = nx.Graph()
    graph.add_nodes_from([host.name for host in hosts])

    for link in links:
        if not link.is_point_to_point:
            continue
        edge = tuple(interface.device_name for interface in link.interfaces)
        for i, interface in enumerate(link.interfaces):
            edge_labels[i][edge] = interface.short_name
        graph.add_edge(*edge)
    return graph, edge_labels


def draw_and_save_topology(graph, edge_labels, target_site):
    plt.figure(figsize=(17, 11))
    pos = nx.spring_layout(graph)
    nx.draw_networkx(graph, pos, node_color='orange')
    nx.draw_networkx_edge_labels(
        graph, pos, edge_labels=edge_labels[0], label_pos=0.8)
    nx.draw_networkx_edge_labels(
        graph, pos, edge_labels=edge_labels[1], label_pos=0.2)
    filename = f"{target_site}_topology.png"
    plt.savefig(filename, format='png')


def device_discovery(subnets_list, site, nr_filtered, ports="T:22,23", arguments="-A --osscan-limit"):
    """ INCOMPLETE AS OF 3/11/2020
    Arguments as string "-sn -n", "--script smb-os-discovery.nse -p 445,137,139"
    can only be run on ios or devices that support textfsm """

    start_time = time.time()
    # devices = []
    discovered_network_devices = []
    # transport_types = ['ssh']
    network_device_identifiers = ["switch", "router", "cisco"]
    for subnet in subnets_list:
        print(f"Scanning subnet: {subnet}")
        nm = nmap.PortScanner()
        nm.scan(hosts=subnet, ports=ports, arguments=arguments)
        for host in nm.all_hosts():

            print(f"Host : {host} hostname: ({nm[host].hostname()})")
            print(f"State : {nm[host].state()}")
            for protocol in nm[host].all_protocols():
                print("----------")
                print(f"Protocol : {protocol}")
                ports = nm[host][protocol].keys()
                discovered_network_device = {}
                for port in ports:
                    print(
                        f"port : {port}, "
                        f"state : {nm[host][protocol][port]['state']}, "
                        f"product: {nm[host][protocol][port]['product']}"
                    )
                    # print(f"All : {nm[host][protocol][port]} ")
                    if any(
                        x in nm[host][protocol][port]["product"].lower()
                        for x in network_device_identifiers
                    ):
                        discovered_network_device["ip"] = host
                        if "ports" in discovered_network_device.keys():
                            discovered_network_device["ports"].append(port)
                        else:
                            discovered_network_device["ports"] = [port]
                        if nm[host].hostname():
                            discovered_network_device["hostname"] = nm[host].hostname(
                            )
                        else:
                            discovered_network_device["hostname"] = host
                        discovered_network_devices.append(
                            discovered_network_device)

            # try to find hostname
            if 'hostscript' in nm[host].keys():
                script_output_lines = [
                    x.strip() for x in nm[host]['hostscript'][0]['output'].splitlines()]
                for line in script_output_lines:
                    # with open('outlines.txt', 'a+') as f:
                    #     f.write(line + '\n')
                    if line.startswith("Computer name:"):
                        hostname = extract_hostname_from_fqdn(
                            line.lstrip("Computer name: "))
                    elif line.startswith("OS:"):
                        os = line.lstrip("OS: ")
            if hostname is None and nm[host]['hostnames'][0]['name'] != '':
                pass

    for device in discovered_network_devices:
        inventory_search_result = nr_filtered.filter(hostname=device["ip"])
        if bool(inventory_search_result):
            print(f"{device['ip']} is already in inventory {inventory_search_result.inventory.hosts} ")
                
        else:
            print(f"{device['ip']} not found in inventory")

    print(f'device discovery found {len(discovered_network_devices)} devices')

    # Time Stuff
    milestone = time.time()
    time_to_run = milestone - start_time
    print(f"{Fore.RED}It took {time_to_run:.2f} seconds for nmap scan{Fore.RESET}")

    # Run
    # nr_filtered.run(task=get_cdp_neighbors)

    # Time Stuff
    milestone = time.time()
    time_to_run = milestone - start_time
    print(f"{Fore.RED}It took {time_to_run:.2f} seconds for get_cdp_neighbors {Fore.RESET}")

    # Run
    # nr_filtered.run(task=populate_arp_table, devices=devices)

    # Time Stuff
    milestone = time.time()
    time_to_run = milestone - start_time
    print(f"{Fore.RED}It took {time_to_run:.2f} seconds for populate_arp_table {Fore.RESET}")

    hosts = nr_filtered.inventory.hosts.values()

    # Run
    # device_int_relationship(hosts=hosts, devices=devices)

    # Time Stuff
    milestone = time.time()
    time_to_run = milestone - start_time
    print(f"{Fore.RED}It took {time_to_run:.2f} seconds for device_int_relationship {Fore.RESET}")

    write_hosts_data_to_inventory_file(hosts=hosts)
